import io
from typing import List

import openpyxl


def extract_excel_contents(
    file_bytes: bytes,
    max_rows_per_sheet: int = 200,
    max_sheets: int = 20,
) -> str:
    """Extracts a compact textual representation across workbook sheets.

    Kept intentionally small and deterministic so it's suitable for sending to
    an LLM. Converts None to empty strings and joins columns with tabs.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    rows: List[str] = []

    for sheet_index, sheet_name in enumerate(wb.sheetnames):
        if sheet_index >= max_sheets:
            rows.append("(...additional sheets truncated...)")
            break

        rows.append(f"=== Sheet: {sheet_name} ===")
        sheet = wb[sheet_name]
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= max_rows_per_sheet:
                rows.append("(...truncated...)")
                break
            rows.append("\t".join(["" if c is None else str(c) for c in row]))
        rows.append("")

    wb.close()

    return "\n".join(rows).strip()


def extract_csv_contents(file_bytes: bytes, max_rows: int = 200) -> str:
    """Extracts a compact textual representation of a CSV file.

    Mirrors the behavior of `extract_excel_contents`: returns up to `max_rows`
    lines joined by newlines, with columns joined by tabs to keep the
    representation compact and consistent for the agent.
    """
    import csv

    text = file_bytes.decode("utf-8", errors="replace")
    rows = []
    reader = csv.reader(text.splitlines())
    for i, row in enumerate(reader):
        if i >= max_rows:
            rows.append("(...truncated...)")
            break
        # convert None/empty to empty string and join with tabs
        rows.append("\t".join(["" if c is None else str(c) for c in row]))

    return "\n".join(rows)
