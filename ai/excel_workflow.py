import os
import io
import base64
from typing import List, Optional, Dict, Any
import zipfile

import httpx
import openpyxl
from agent_framework import (
    Executor,
    WorkflowBuilder,
    WorkflowContext,
    executor,
    handler,
)
from typing_extensions import Never
from dotenv import load_dotenv

load_dotenv()

from agent_framework.openai import OpenAIChatClient


def extract_excel_contents(file_bytes: bytes, max_rows: int = 200) -> str:
    """Extracts a textual representation of the first worksheet from an .xlsx file.

    Returns a compact, newline-separated representation suitable for sending to an LLM.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = wb[wb.sheetnames[0]]

    rows = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= max_rows:
            rows.append("(...truncated...)")
            break
        # convert None to empty string and join with tabs for readability
        rows.append("\t".join(["" if c is None else str(c) for c in row]))

    return "\n".join(rows)


async def convert_excel_to_pdf_collabora(
    file_bytes: bytes,
    collabora_base_url: str = "http://localhost:8080",
    timeout: int = 60,
) -> bytes:
    """Sends the excel file to a Collabora (CODE/LOOL) server convert endpoint and returns PDF bytes.

    collabora_base_url should include scheme and host (and port). The function posts the file as
    multipart form data to the common convert endpoint: /lool/convert-to/pdf
    """
    convert_url = collabora_base_url.rstrip("/") + "/lool/convert-to/pdf"

    files = {"file": ("file.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}

    async with httpx.AsyncClient(timeout=timeout) as client:
        resp = await client.post(convert_url, files=files)
        resp.raise_for_status()
        return resp.content


async def convert_pdf_to_png_collabora(
    pdf_bytes: bytes,
    collabora_base_url: str = "http://localhost:8080",
    timeout: int = 60,
) -> List[bytes]:
    """Send a PDF to Collabora convert endpoint asking for PNG(s).

    Collabora may return a single PNG image (image/png) or a ZIP archive of PNGs
    (application/zip) when converting multi-page PDFs. This function detects both
    and returns a list of PNG bytes.
    """
    convert_url = collabora_base_url.rstrip("/") + "/lool/convert-to/png"

    files = {"file": ("file.pdf", pdf_bytes, "application/pdf")}

    async with httpx.AsyncClient(timeout=timeout) as client:
        resp = await client.post(convert_url, files=files)
        resp.raise_for_status()
        content = resp.content
        # If it's a zip, extract png files
        try:
            b = io.BytesIO(content)
            if zipfile.is_zipfile(b):
                pngs: List[bytes] = []
                with zipfile.ZipFile(b) as z:
                    for name in z.namelist():
                        if name.lower().endswith(".png"):
                            pngs.append(z.read(name))
                return pngs
        except Exception:
            # fall through and treat as single image
            pass

        # Not a ZIP — assume the response body is a single image (PNG)
        return [content]


async def run_excel_agent_workflow(
    excel_bytes: bytes,
    collabora_base_url: Optional[str] = None,
    agent_prompt: str = "Please analyze the spreadsheet and the associated image(s).",
    model_env: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """End-to-end workflow: extract, convert, rasterize, and call agent with both inputs.

    Returns a dict containing the agent text, the extracted text, and base64 PNG previews.
    """
    # We'll build a small Workflow using agent_framework.WorkflowBuilder.
    # Design:
    #  - file_executor: receives raw excel bytes and forwards the bytes to both
    #    extract_executor and convert_executor.
    #  - extract_executor: extracts text and sends a dict {"extracted": text}
    #  - convert_executor -> pdf_executor -> png_executor: produces a base64 png
    #    and sends a dict {"png_b64": "..."}
    #  - agent_collector: collects messages from extract and png executors; when it
    #    has both, it calls the agent and yields a single workflow output (the agent result)

    # Simple start executor — forwards the incoming bytes to downstream nodes.
    @executor(id="file_executor")
    async def file_executor(data: bytes, ctx: WorkflowContext[bytes]) -> None:
        await ctx.send_message(data)

    # Extract executor — converts bytes -> extracted text and forwards a dict
    @executor(id="extract_executor")
    async def extract_executor(data: bytes, ctx: WorkflowContext[dict]) -> None:
        text = extract_excel_contents(data)
        await ctx.send_message({"extracted": text})

    # Convert executor chain: excel bytes -> pdf -> png (base64)
    @executor(id="convert_to_pdf_executor")
    async def convert_to_pdf_executor(data: bytes, ctx: WorkflowContext[bytes]) -> None:
        collabora = collabora_base_url or os.getenv("COLLABORA_URL", "http://localhost:8080")
        pdf = await convert_excel_to_pdf_collabora(data, collabora_base_url=collabora)
        await ctx.send_message(pdf)

    @executor(id="pdf_to_png_executor")
    async def pdf_to_png_executor(pdf_bytes: bytes, ctx: WorkflowContext[dict]) -> None:
        collabora = collabora_base_url or os.getenv("COLLABORA_URL", "http://localhost:8080")
        pngs = await convert_pdf_to_png_collabora(pdf_bytes, collabora_base_url=collabora)
        # Use the first page as preview and base64 encode it
        first_b64 = base64.b64encode(pngs[0]).decode("ascii") if pngs else ""
        await ctx.send_message({"png_b64": first_b64})

    # Agent-collector executor: accumulate partial inputs and call the agent
    class AgentCollector(Executor):
        def __init__(self, id: str):
            super().__init__(id=id)
            # naive in-memory buffer keyed by a single run; for demo only
            self._buffer: Dict[str, Any] = {}

        @handler
        async def handle(self, message: dict, ctx: WorkflowContext[Never, dict]) -> None:
            # Merge incoming dict into buffer
            self._buffer.update(message)

            # If we have both pieces, run the agent and yield the output
            if "extracted" in self._buffer and "png_b64" in self._buffer:
                extracted = self._buffer.pop("extracted")
                png_b64 = self._buffer.pop("png_b64")

                # Prepare agent client
                api_key = os.getenv("OLLAMA_API_KEY")
                if not api_key:
                    raise RuntimeError("OLLAMA_API_KEY required to run agent")

                base_url = os.getenv("OLLAMA_CLOUD_URL", "http://localhost:11434/v1/")
                model_id = os.getenv("OLLAMA_MODEL_ID", "deepseek-r1:8b")

                client = OpenAIChatClient(api_key=api_key, base_url=base_url, model_id=model_id)

                instructions = (
                    "You will be given two inputs:\n"
                    "1) A textual extraction of an Excel spreadsheet.\n"
                    "2) A PNG image rendering of the spreadsheet (base64).\n"
                    "Use both sources for extraction, validation, calculations, and produce a concise JSON report."
                )

                agent = client.create_agent(name="excel_inspector", instructions=instructions)

                full_prompt = (
                    f"User prompt: {agent_prompt}\n\n"
                    "---EXTRACTED_SPREADSHEET_TEXT---\n"
                    f"{extracted}\n\n"
                    "---PNG_BASE64_FIRST_PAGE---\n"
                    f"{png_b64}\n"
                    "---END---\n"
                    "When giving results, output a JSON object with keys: summary, tables (if any), calculations (if requested)."
                )

                result = await agent.run(full_prompt)

                # Yield the result as the workflow output; downstream consumers can read outputs
                await ctx.yield_output({"agent_response": str(result), "extracted_text": extracted, "png_b64": png_b64})

    agent_collector = AgentCollector(id="agent_collector")

    # Build the workflow graph:
    # file_executor -> extract_executor -> agent_collector
    # file_executor -> convert_to_pdf_executor -> pdf_to_png_executor -> agent_collector
    workflow = (
        WorkflowBuilder()
        .set_start_executor(file_executor)
        .add_edge(file_executor, extract_executor)
        .add_edge(file_executor, convert_to_pdf_executor)
        .add_edge(extract_executor, agent_collector)
        .add_edge(convert_to_pdf_executor, pdf_to_png_executor)
        .add_edge(pdf_to_png_executor, agent_collector)
        .build()
    )

    # Run workflow and gather outputs
    events = await workflow.run(excel_bytes)

    outputs = events.get_outputs() or []
    # Return last output or empty structure
    if outputs:
        out = outputs[-1]
        return {
            "agent_response": out.get("agent_response"),
            "extracted_text": out.get("extracted_text"),
            "pngs_base64": [out.get("png_b64")],
        }

    return {"agent_response": "", "extracted_text": "", "pngs_base64": []}
