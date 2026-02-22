import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Protection
import csv
from .models import ProductInput, ProductsList


# Shopify CSV template headers (taken from provided product_template.csv)
TEMPLATE_HEADERS = [
    "Handle",
    "Title",
    "Body (HTML)",
    "Vendor",
    "Product Category",
    "Type",
    "Tags",
    "Published",
    "Option1 Name",
    "Option1 Value",
    "Option2 Name",
    "Option2 Value",
    "Option3 Name",
    "Option3 Value",
    "Variant SKU",
    "Variant Grams",
    "Variant Inventory Tracker",
    "Variant Inventory Qty",
    "Variant Inventory Policy",
    "Variant Fulfillment Service",
    "Variant Price",
    "Variant Compare At Price",
    "Variant Requires Shipping",
    "Variant Taxable",
    "Variant Barcode",
    "Image Src",
    "Image Position",
    "Image Alt Text",
    "Gift Card",
    "SEO Title",
    "SEO Description",
    "Google Shopping / Google Product Category",
    "Google Shopping / Gender",
    "Google Shopping / Age Group",
    "Google Shopping / MPN",
    "Google Shopping / AdWords Grouping",
    "Google Shopping / AdWords Labels",
    "Google Shopping / Condition",
    "Google Shopping / Custom Product",
    "Google Shopping / Custom Label 0",
    "Google Shopping / Custom Label 1",
    "Google Shopping / Custom Label 2",
    "Google Shopping / Custom Label 3",
    "Google Shopping / Custom Label 4",
    "Variant Image",
    "Variant Weight Unit",
    "Variant Tax Code",
    "Cost per item",
    "Price / International",
    "Compare At Price / International",
    "Status",
]


def _slugify_handle(title: str) -> str:
    # simple slug: lowercase, replace spaces with dashes, remove non-url chars
    import re

    s = (title or "").strip().lower()
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^a-z0-9\-]", "", s)
    return s


HEADER_INDEX = {header: index for index, header in enumerate(TEMPLATE_HEADERS)}


def _to_csv_bool(value: bool | str | None, *, default: str = "") -> str:
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "1", "yes", "on"}:
            return "TRUE"
        if normalized in {"false", "0", "no", "off"}:
            return "FALSE"
    return default


def _tags_to_csv(tags: str | list[str] | None) -> str:
    if isinstance(tags, str):
        return tags
    if isinstance(tags, list):
        return ", ".join([str(tag).strip() for tag in tags if str(tag).strip()])
    return ""


def _set_product_level_columns(
    row: list[str],
    *,
    product: ProductInput,
    handle: str,
    title: str,
    body: str,
    vendor: str,
    published: str,
    opt1_name: str,
    opt2_name: str,
    opt3_name: str,
    image_src: str,
    image_position: str,
    image_alt_text: str,
) -> None:
    row[HEADER_INDEX["Handle"]] = handle
    row[HEADER_INDEX["Title"]] = title
    row[HEADER_INDEX["Body (HTML)"]] = body
    row[HEADER_INDEX["Vendor"]] = vendor
    row[HEADER_INDEX["Product Category"]] = product.product_category or ""
    row[HEADER_INDEX["Type"]] = product.product_type or ""
    row[HEADER_INDEX["Tags"]] = _tags_to_csv(product.tags)
    row[HEADER_INDEX["Published"]] = published
    row[HEADER_INDEX["Option1 Name"]] = opt1_name
    if opt2_name:
        row[HEADER_INDEX["Option2 Name"]] = opt2_name
    if opt3_name:
        row[HEADER_INDEX["Option3 Name"]] = opt3_name
    row[HEADER_INDEX["Image Src"]] = image_src
    row[HEADER_INDEX["Image Position"]] = image_position
    row[HEADER_INDEX["Image Alt Text"]] = image_alt_text
    row[HEADER_INDEX["Gift Card"]] = _to_csv_bool(product.gift_card)
    row[HEADER_INDEX["SEO Title"]] = product.seo_title or ""
    row[HEADER_INDEX["SEO Description"]] = product.seo_description or ""
    row[HEADER_INDEX["Google Shopping / Google Product Category"]] = (
        product.google_shopping_category or ""
    )
    row[HEADER_INDEX["Google Shopping / Gender"]] = product.google_shopping_gender or ""
    row[HEADER_INDEX["Google Shopping / Age Group"]] = (
        product.google_shopping_age_group or ""
    )
    row[HEADER_INDEX["Google Shopping / MPN"]] = product.google_shopping_mpn or ""
    row[HEADER_INDEX["Google Shopping / AdWords Grouping"]] = (
        product.google_adwords_grouping or ""
    )
    row[HEADER_INDEX["Google Shopping / AdWords Labels"]] = (
        product.google_adwords_labels or ""
    )
    row[HEADER_INDEX["Google Shopping / Condition"]] = (
        product.google_shopping_condition or ""
    )
    row[HEADER_INDEX["Google Shopping / Custom Product"]] = _to_csv_bool(
        product.google_shopping_custom_product
    )
    row[HEADER_INDEX["Google Shopping / Custom Label 0"]] = (
        product.google_custom_label_0 or ""
    )
    row[HEADER_INDEX["Google Shopping / Custom Label 1"]] = (
        product.google_custom_label_1 or ""
    )
    row[HEADER_INDEX["Google Shopping / Custom Label 2"]] = (
        product.google_custom_label_2 or ""
    )
    row[HEADER_INDEX["Google Shopping / Custom Label 3"]] = (
        product.google_custom_label_3 or ""
    )
    row[HEADER_INDEX["Google Shopping / Custom Label 4"]] = (
        product.google_custom_label_4 or ""
    )
    row[HEADER_INDEX["Status"]] = product.status or "active"


def _product_to_rows(product: ProductInput) -> list[list[str]]:
    """Serialize a ProductInput into one or more CSV rows following Shopify format.

    For products with multiple variants, the first row contains product-level
    information and the first variant; subsequent rows contain only Handle and
    variant-specific fields as per Shopify import CSV conventions.
    """
    rows: list[list[str]] = []

    handle = (product.handle or "").strip() or _slugify_handle(product.title)
    title = product.title
    body = product.body_html or ""
    vendor = product.vendor or ""
    published = _to_csv_bool(product.published, default="TRUE")

    options = product.options or []
    opt1_name = options[0].name if len(options) >= 1 else "Option1"
    opt2_name = options[1].name if len(options) >= 2 else ""
    opt3_name = options[2].name if len(options) >= 3 else ""

    images = product.images or []
    first_image = product.image_src or (images[0].src if images else "") or ""
    first_image_position = (
        str(product.image_position)
        if isinstance(product.image_position, int)
        else (
            str(images[0].position)
            if images and isinstance(images[0].position, int)
            else ""
        )
    )
    first_image_alt = (
        product.image_alt_text
        or (images[0].alt if images and images[0].alt else "")
        or ""
    )

    variants = product.variants or []
    if not variants:
        # Create a single default variant row
        row = ["" for _ in TEMPLATE_HEADERS]
        _set_product_level_columns(
            row,
            product=product,
            handle=handle,
            title=title,
            body=body,
            vendor=vendor,
            published=published,
            opt1_name=opt1_name,
            opt2_name=opt2_name,
            opt3_name=opt3_name,
            image_src=first_image,
            image_position=first_image_position,
            image_alt_text=first_image_alt,
        )
        row[HEADER_INDEX["Option1 Value"]] = "Default"
        row[HEADER_INDEX["Variant Weight Unit"]] = "g"
        rows.append(row)
        return rows

    # For products with variants, write first row with product-level and first variant
    for idx, variant in enumerate(variants):
        row = ["" for _ in TEMPLATE_HEADERS]
        row[HEADER_INDEX["Handle"]] = handle
        if idx == 0:
            _set_product_level_columns(
                row,
                product=product,
                handle=handle,
                title=title,
                body=body,
                vendor=vendor,
                published=published,
                opt1_name=opt1_name,
                opt2_name=opt2_name,
                opt3_name=opt3_name,
                image_src=first_image,
                image_position=first_image_position,
                image_alt_text=first_image_alt,
            )

        # Variant-specific fields
        if variant.sku:
            row[HEADER_INDEX["Variant SKU"]] = variant.sku
        if variant.grams is not None:
            row[HEADER_INDEX["Variant Grams"]] = str(variant.grams)
        if variant.inventory_tracker:
            row[HEADER_INDEX["Variant Inventory Tracker"]] = variant.inventory_tracker
        if variant.price is not None:
            row[HEADER_INDEX["Variant Price"]] = str(variant.price)
        if variant.compare_at_price is not None:
            row[HEADER_INDEX["Variant Compare At Price"]] = str(variant.compare_at_price)
        if variant.inventory_quantity is not None:
            row[HEADER_INDEX["Variant Inventory Qty"]] = str(
                variant.inventory_quantity
            )
        if variant.inventory_policy:
            row[HEADER_INDEX["Variant Inventory Policy"]] = variant.inventory_policy
        if variant.fulfillment_service:
            row[HEADER_INDEX["Variant Fulfillment Service"]] = variant.fulfillment_service
        if variant.requires_shipping is not None:
            row[HEADER_INDEX["Variant Requires Shipping"]] = _to_csv_bool(
                variant.requires_shipping
            )
        if variant.taxable is not None:
            row[HEADER_INDEX["Variant Taxable"]] = _to_csv_bool(variant.taxable)
        if variant.barcode:
            row[HEADER_INDEX["Variant Barcode"]] = variant.barcode
        if variant.variant_image:
            row[HEADER_INDEX["Variant Image"]] = variant.variant_image
        if variant.variant_weight_unit:
            row[HEADER_INDEX["Variant Weight Unit"]] = variant.variant_weight_unit
        if variant.variant_tax_code:
            row[HEADER_INDEX["Variant Tax Code"]] = variant.variant_tax_code
        if variant.cost_per_item is not None:
            row[HEADER_INDEX["Cost per item"]] = str(variant.cost_per_item)
        if variant.price_international:
            row[HEADER_INDEX["Price / International"]] = variant.price_international
        if variant.compare_at_price_international:
            row[HEADER_INDEX["Compare At Price / International"]] = (
                variant.compare_at_price_international
            )
        if variant.status:
            row[HEADER_INDEX["Status"]] = variant.status
        # Option values
        if variant.option1:
            row[HEADER_INDEX["Option1 Value"]] = variant.option1
        if variant.option2:
            row[HEADER_INDEX["Option2 Value"]] = variant.option2
        if variant.option3:
            row[HEADER_INDEX["Option3 Value"]] = variant.option3

        # Defaults
        if not row[HEADER_INDEX["Variant Inventory Policy"]]:
            row[HEADER_INDEX["Variant Inventory Policy"]] = "deny"
        if not row[HEADER_INDEX["Variant Fulfillment Service"]]:
            row[HEADER_INDEX["Variant Fulfillment Service"]] = "manual"
        if not row[HEADER_INDEX["Variant Requires Shipping"]]:
            row[HEADER_INDEX["Variant Requires Shipping"]] = "TRUE"
        if not row[HEADER_INDEX["Variant Taxable"]]:
            row[HEADER_INDEX["Variant Taxable"]] = "TRUE"
        if not row[HEADER_INDEX["Variant Weight Unit"]]:
            row[HEADER_INDEX["Variant Weight Unit"]] = "g"

        rows.append(row)

    return rows


def _append_ai_enhancements_sheet(workbook: Workbook, products: list[ProductInput]) -> None:
    worksheet = workbook.create_sheet("AI Enhancements")
    worksheet.append(["Handle", "Title", "Attribute", "Value", "Type"])
    wrote_rows = False
    for product in products:
        handle = (product.handle or "").strip() or _slugify_handle(product.title)
        title = product.title
        for metafield in product.metafields or []:
            worksheet.append(
                [
                    handle,
                    title,
                    f"metafield:{metafield.namespace}.{metafield.key}",
                    metafield.value,
                    metafield.type or "single_line_text_field",
                ]
            )
            wrote_rows = True
    if not wrote_rows:
        worksheet.append(["", "", "No AI-only attributes captured", "", ""])
    _protect_header_row(worksheet)


def _protect_header_row(worksheet) -> None:
    """Lock row 1 headers and keep data rows editable."""
    for cell in worksheet[1]:
        cell.protection = Protection(locked=True)
    for row in worksheet.iter_rows(
        min_row=2,
        max_row=max(worksheet.max_row, 2),
        min_col=1,
        max_col=max(worksheet.max_column, 1),
    ):
        for cell in row:
            cell.protection = Protection(locked=False)
    worksheet.protection.sheet = True


def create_excel_workbook(products_list: ProductsList, output_path: str) -> str:
    """Create an XLSX workbook (Shopify template columns) for Collabora preview."""
    if not output_path:
        raise ValueError("output_path must be provided")

    absolute_path = os.path.abspath(output_path)
    target_dir = os.path.dirname(absolute_path)
    if target_dir and not os.path.exists(target_dir):
        os.makedirs(target_dir, exist_ok=True)

    xlsx_path = (
        absolute_path
        if absolute_path.lower().endswith(".xlsx")
        else f"{absolute_path}.xlsx"
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(TEMPLATE_HEADERS)
    for product in products_list.products:
        for row in _product_to_rows(product):
            ws.append(row)
    _protect_header_row(ws)
    _append_ai_enhancements_sheet(wb, products_list.products)
    wb.save(xlsx_path)
    return xlsx_path


def create_excel_bytes(products_list: ProductsList) -> bytes:
    """Return XLSX bytes for the given ProductsList without writing to disk."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(TEMPLATE_HEADERS)
    for product in products_list.products:
        for row in _product_to_rows(product):
            ws.append(row)
    _protect_header_row(ws)
    _append_ai_enhancements_sheet(wb, products_list.products)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_csv_bytes(products_list: ProductsList) -> bytes:
    """Return CSV bytes for the given ProductsList without writing to disk."""
    import io
    import csv as _csv

    output = io.StringIO()
    writer = _csv.writer(output)
    writer.writerow(TEMPLATE_HEADERS)
    for product in products_list.products:
        rows = _product_to_rows(product)
        for r in rows:
            writer.writerow(r)

    return output.getvalue().encode("utf-8")
