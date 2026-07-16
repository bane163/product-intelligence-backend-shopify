"""Conservatively add exact review-field coordinates to spreadsheet products."""

from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


ALIASES = {
    "title": {"title", "product title", "product name", "name"},
    "vendor": {"vendor", "brand", "manufacturer", "supplier"},
    "sku": {"sku", "variant sku", "product code", "item code"},
    "price": {"price", "variant price", "unit price", "retail price"},
}


def _text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _price(value: Any) -> Decimal | None:
    cleaned = re.sub(r"[^0-9.\-]", "", str(value or "").replace(",", ""))
    try:
        return Decimal(cleaned).normalize() if cleaned else None
    except InvalidOperation:
        return None


def _matches(field: str, left: Any, right: Any) -> bool:
    if field == "price":
        a, b = _price(left), _price(right)
        return a is not None and b is not None and a == b
    return bool(_text(left)) and _text(left) == _text(right)


def _canonical_ref_field(value: Any) -> str | None:
    field = re.sub(r"\[(\d+)\]", r".\1", str(value or "").strip().casefold())
    field = re.sub(r"[_\s-]+", ".", field)
    if field in {"title", "product.title", "product.name", "name"}:
        return "title"
    if field in {"vendor", "product.vendor", "brand", "manufacturer", "supplier"}:
        return "vendor"
    if re.fullmatch(r"(?:product\.)?sku|variants?(?:\.\d+)?\.sku", field):
        return "sku"
    if re.fullmatch(r"(?:product\.)?price|variants?(?:\.\d+)?\.price", field):
        return "price"
    return None


def _is_variants_ref(value: Any) -> bool:
    field = re.sub(r"\[(\d+)\]", r".\1", str(value or "").strip().casefold())
    field = re.sub(r"[_\s-]+", ".", field)
    return bool(re.fullmatch(r"variants?(?:\.\d+)?", field))


def _split_qualified_ref(value: Any) -> tuple[str | None, str]:
    """Split Sheet!A1, including Excel's quoted/apostrophe sheet syntax."""
    raw = str(value or "").strip()
    if "!" not in raw:
        return None, raw
    sheet, coordinate = raw.rsplit("!", 1)
    sheet = sheet.strip()
    if len(sheet) >= 2 and sheet[0] == sheet[-1] == "'":
        sheet = sheet[1:-1].replace("''", "'")
    return sheet, coordinate.strip()


def _ref_location(ref: dict[str, Any]) -> tuple[str | None, str]:
    raw = ref.get("cell") or ref.get("cell_range") or ""
    qualified_sheet, coordinate = _split_qualified_ref(raw)
    declared_sheet = str(ref.get("sheet") or "").strip() or None
    if qualified_sheet and declared_sheet and qualified_sheet != declared_sheet:
        return None, ""
    return qualified_sheet or declared_sheet, coordinate


def _first_row(coordinate: str) -> int | None:
    try:
        return range_boundaries(coordinate)[1]
    except (TypeError, ValueError):
        return None


def _rows(file_bytes: bytes, filename: str | None) -> Iterable[tuple[str, list[list[Any]]]]:
    if (filename or "").lower().endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        yield "Sheet1", [list(row) for row in csv.reader(io.StringIO(text))]
        return
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            yield sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _table_for_sheet(
    tables: list[tuple[str, list[list[Any]], int, dict[str, int]]], sheet: str
) -> tuple[str, list[list[Any]], int, dict[str, int]] | None:
    return next((table for table in tables if table[0] == sheet), None)


def enrich_spreadsheet_source_refs(
    products: list[dict[str, Any]], file_bytes: bytes, filename: str | None
) -> None:
    """Mutate products with deterministic, sheet-qualified single-cell refs."""
    tables: list[tuple[str, list[list[Any]], int, dict[str, int]]] = []
    try:
        sheet_rows = list(_rows(file_bytes, filename))
    except Exception:
        return
    for sheet, rows in sheet_rows:
        for header_index, row in enumerate(rows[:20]):
            columns: dict[str, int] = {}
            for column, value in enumerate(row):
                normalized = _text(value)
                for field, aliases in ALIASES.items():
                    if normalized in aliases and field not in columns:
                        columns[field] = column
            if "title" in columns or "sku" in columns:
                tables.append((sheet, rows, header_index, columns))
                break

    for product in products:
        refs = product.get("source_refs")
        if not isinstance(refs, list):
            refs = []
            product["source_refs"] = refs
        dict_refs = [ref for ref in refs if isinstance(ref, dict)]
        variants = product.get("variants") if isinstance(product.get("variants"), list) else []
        first_variant = variants[0] if variants and isinstance(variants[0], dict) else {}
        values = {
            "title": product.get("title"),
            "vendor": product.get("vendor"),
            "sku": first_variant.get("sku"),
            "price": first_variant.get("price"),
        }

        # A variants range or an exact title reference anchors this product table.
        anchors: list[tuple[str, int]] = []
        for ref in dict_refs:
            canonical = _canonical_ref_field(ref.get("field"))
            if not (_is_variants_ref(ref.get("field")) or (canonical == "title" and ref.get("cell"))):
                continue
            sheet, coordinate = _ref_location(ref)
            row_index = _first_row(coordinate)
            if sheet and row_index:
                anchors.append((sheet, row_index - 1))

        anchored_sheets = {sheet for sheet, _ in anchors}
        anchored_rows = {row for _, row in anchors}
        candidate: tuple[str, list[list[Any]], int, dict[str, int], int] | None = None
        anchors_conflict = len(anchored_sheets) > 1 or len(anchored_rows) > 1
        if anchors and not anchors_conflict:
            sheet, row_index = anchors[0]
            table = _table_for_sheet(tables, sheet)
            if table and table[2] < row_index < len(table[1]):
                candidate = (*table, row_index)
        elif not anchors:
            candidates: list[tuple[str, list[list[Any]], int, dict[str, int], int]] = []
            for table in tables:
                sheet, rows, header_index, columns = table
                data_indexes = range(header_index + 1, len(rows))
                sku_matches = [
                    i for i in data_indexes
                    if "sku" in columns
                    and _matches("sku", rows[i][columns["sku"]] if columns["sku"] < len(rows[i]) else None, values["sku"])
                ]
                if values["sku"] is not None and len(sku_matches) == 1:
                    candidates.append((*table, sku_matches[0]))
                    continue
                title_matches = [
                    i for i in data_indexes
                    if "title" in columns
                    and _matches("title", rows[i][columns["title"]] if columns["title"] < len(rows[i]) else None, values["title"])
                ]
                if len(title_matches) == 1:
                    candidates.append((*table, title_matches[0]))
            if len(candidates) == 1:
                candidate = candidates[0]

        # SKU refs are usable only when they name a real SKU cell with the expected value.
        valid_sku_ref: dict[str, Any] | None = None
        for ref in dict_refs:
            if _canonical_ref_field(ref.get("field")) != "sku":
                continue
            sheet, coordinate = _ref_location(ref)
            table = _table_for_sheet(tables, sheet) if sheet else None
            try:
                min_col, min_row, max_col, max_row = range_boundaries(coordinate)
            except (TypeError, ValueError):
                continue
            if (
                table
                and min_col == max_col
                and min_row == max_row
                and table[3].get("sku") == min_col - 1
                and min_row <= len(table[1])
                and min_col <= len(table[1][min_row - 1])
                and _matches("sku", table[1][min_row - 1][min_col - 1], values["sku"])
                and (not anchored_sheets or sheet in anchored_sheets)
            ):
                ref["sheet"] = sheet
                ref["cell"] = coordinate
                ref.pop("cell_range", None)
                valid_sku_ref = ref
                break
        refs[:] = [
            ref for ref in refs
            if not isinstance(ref, dict)
            or _canonical_ref_field(ref.get("field")) != "sku"
            or ref is valid_sku_ref
        ]

        if not candidate:
            continue
        sheet, rows, _, columns, row_index = candidate
        row = rows[row_index]
        present = {
            canonical
            for ref in refs
            if isinstance(ref, dict)
            for canonical in [_canonical_ref_field(ref.get("field"))]
            if canonical and (ref.get("cell") or ref.get("cell_range"))
        }
        for field in ("title", "vendor", "sku", "price"):
            if field in present or field not in columns or values[field] is None:
                continue
            column = columns[field]
            cell_value = row[column] if column < len(row) else None
            if not _matches(field, cell_value, values[field]):
                continue
            coordinate = f"{get_column_letter(column + 1)}{row_index + 1}"
            refs.append({
                "field": field,
                "document_kind": "spreadsheet",
                "sheet": sheet,
                "cell": coordinate,
                "value": str(cell_value),
            })
