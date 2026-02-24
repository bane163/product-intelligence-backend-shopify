"""Use-case: create highlighted source artifacts for source verification."""

from __future__ import annotations

import io
import os
import re
import uuid
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

from application.ports.collabora_port import CollaboraPort
from application.ports.supabase_port import SupabaseNamespacedPort
from application.services.document_formats import classify_document

_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLSM_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroenabled.12"
_PDF_CONTENT_TYPE = "application/pdf"
_HIGHLIGHT_FILL = PatternFill(
    fill_type="solid",
    start_color="FFF59D",
    end_color="FFF59D",
)


def _optional_str(data: dict[str, Any], key: str) -> str | None:
    value = data.get(key)
    return value if isinstance(value, str) and value.strip() else None


def _normalize_sheet_name(value: str) -> str:
    trimmed = value.strip()
    if trimmed.startswith("'") and trimmed.endswith("'") and len(trimmed) >= 2:
        return trimmed[1:-1].replace("''", "'")
    return trimmed


def _split_sheet_and_ref(value: str) -> tuple[str | None, str]:
    trimmed = value.strip()
    if not trimmed:
        return None, ""

    if "!" in trimmed:
        raw_sheet, raw_ref = trimmed.rsplit("!", 1)
        return _normalize_sheet_name(raw_sheet), raw_ref.strip()

    quoted_dot_match = re.match(r"^('(?:''|[^'])+')\.(.+)$", trimmed)
    if quoted_dot_match:
        return _normalize_sheet_name(quoted_dot_match.group(1)), quoted_dot_match.group(
            2
        ).strip()

    plain_dot_match = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\.(.+)$", trimmed)
    if plain_dot_match:
        return plain_dot_match.group(1).strip(), plain_dot_match.group(2).strip()

    return None, trimmed


def _normalize_cell_ref(value: str) -> str:
    normalized = value.replace("$", "").strip().upper()
    match = re.match(r"^([A-Z]+)(\d+)$", normalized)
    if not match:
        raise ValueError(f"Invalid cell reference: {value}")
    row = int(match.group(2))
    if row < 1:
        raise ValueError(f"Invalid row index in cell reference: {value}")
    return f"{match.group(1)}{row}"


def _resolve_target_range(worksheet, target_ref: str) -> str:
    normalized = target_ref.replace("$", "").strip().upper()
    if not normalized:
        raise ValueError("Missing source coordinates")

    max_col = max(worksheet.max_column or 0, 1)

    row_range_match = re.match(r"^(\d+)(?::(\d+))?$", normalized)
    if row_range_match:
        start_row = int(row_range_match.group(1))
        end_row = int(row_range_match.group(2) or row_range_match.group(1))
        if start_row < 1 or end_row < 1:
            raise ValueError(f"Invalid row range: {target_ref}")
        if start_row > end_row:
            start_row, end_row = end_row, start_row
        return f"A{start_row}:{get_column_letter(max_col)}{end_row}"

    if ":" in normalized:
        left_raw, right_raw = normalized.split(":", 1)
        left = _normalize_cell_ref(left_raw)
        right = _normalize_cell_ref(right_raw)
        min_col, min_row, max_col_idx, max_row = range_boundaries(f"{left}:{right}")
        return (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col_idx)}{max_row}"
        )

    return _normalize_cell_ref(normalized)


def _build_target(
    *,
    workbook,
    raw_sheet: str | None,
    raw_ref: str,
    fallback_sheet: str | None,
    field: str | None = None,
) -> dict[str, str] | None:
    source_sheet, source_ref = _split_sheet_and_ref(raw_ref)
    normalized_sheet = _normalize_sheet_name(
        source_sheet or (raw_sheet or "").strip() or (fallback_sheet or "").strip()
    )
    target_sheet = normalized_sheet or workbook.active.title
    if target_sheet not in workbook.sheetnames:
        return None

    worksheet = workbook[target_sheet]
    target_ref = source_ref or raw_ref
    normalized_target_range = _resolve_target_range(worksheet, target_ref)
    target: dict[str, str] = {
        "sheet": target_sheet,
        "cell_range": normalized_target_range,
    }
    if field:
        target["field"] = field
    return target


def _optional_positive_int(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float) and value.is_integer():
        int_value = int(value)
        return int_value if int_value > 0 else None
    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return None
        if trimmed.isdigit():
            int_value = int(trimmed)
            return int_value if int_value > 0 else None
    return None


def _optional_trimmed_str(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    trimmed = value.strip()
    return trimmed or None


def _normalize_bbox(value: Any) -> tuple[float, float, float, float] | None:
    if not isinstance(value, (list, tuple)) or len(value) != 4:
        return None
    try:
        x0, y0, x1, y1 = (float(value[0]), float(value[1]), float(value[2]), float(value[3]))
    except (TypeError, ValueError):
        return None
    left, right = sorted((x0, x1))
    top, bottom = sorted((y0, y1))
    if right <= left or bottom <= top:
        return None
    return left, top, right, bottom


_TEXT_FIRST_SOURCE_FIELDS = {
    "title",
    "handle",
    "body_html",
    "vendor",
    "product_type",
    "tags",
    "sku",
    "price",
}
_MEDIA_LAST_SOURCE_FIELDS = {"image_src", "image_alt_text", "image_position"}


def _source_ref_field_priority(field_value: Any) -> int:
    field = _optional_trimmed_str(field_value)
    if not field:
        return 50
    normalized = field.lower()
    if normalized == "title":
        return 0
    if normalized in _TEXT_FIRST_SOURCE_FIELDS:
        return 10
    if (
        normalized.startswith("variant")
        or normalized.startswith("option")
        or "sku" in normalized
        or "price" in normalized
    ):
        return 20
    if normalized in _MEDIA_LAST_SOURCE_FIELDS or "image" in normalized:
        return 90
    return 40


def _select_highlight_page_candidate(
    current: tuple[int, int, int] | None,
    *,
    field: Any,
    target_index: int,
    page: int,
) -> tuple[int, int, int]:
    priority = _source_ref_field_priority(field)
    candidate = (priority, target_index, page)
    if current is None:
        return candidate
    current_priority, current_target_index, _ = current
    if priority < current_priority:
        return candidate
    if priority == current_priority and target_index < current_target_index:
        return candidate
    return current


def _collect_pdf_targets(
    source_refs: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    targets: list[dict[str, Any]] = []
    for source_ref in source_refs:
        if not isinstance(source_ref, dict):
            continue
        page = _optional_positive_int(source_ref.get("page"))
        field = _optional_trimmed_str(source_ref.get("field"))
        bbox = _normalize_bbox(source_ref.get("bbox"))
        value = _optional_trimmed_str(source_ref.get("value"))
        if bbox is not None and page is not None:
            targets.append(
                {
                    "kind": "bbox",
                    "page": page,
                    "bbox": bbox,
                    "field": field.lower() if field else None,
                }
            )
        if value is not None:
            targets.append(
                {
                    "kind": "text",
                    "page": page,
                    "value": value,
                    "field": field.lower() if field else None,
                }
            )
    if not targets:
        raise ValueError(
            "Missing non-spreadsheet source highlight target (expected page+bbox, page+value, or value)"
        )
    return targets


def _highlight_pdf_bytes(
    *,
    pdf_bytes: bytes,
    targets: list[dict[str, Any]],
) -> tuple[bytes, int]:
    try:
        import fitz  # type: ignore
    except Exception as exc:  # pragma: no cover - dependency/runtime guarded
        raise RuntimeError(
            "PyMuPDF dependency is required for non-spreadsheet source highlighting"
        ) from exc

    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        did_highlight = False
        selected_page_candidate: tuple[int, int, int] | None = None
        for target_index, target in enumerate(targets):
            if target["kind"] == "bbox":
                page_number = _optional_positive_int(target.get("page"))
                if page_number is None:
                    continue
                page_index = page_number - 1
                if page_index < 0 or page_index >= document.page_count:
                    continue
                page = document[page_index]
                x0, y0, x1, y1 = target["bbox"]
                if (
                    0 <= x0 <= 1
                    and 0 <= y0 <= 1
                    and 0 <= x1 <= 1
                    and 0 <= y1 <= 1
                ):
                    page_rect = page.rect
                    x0, x1 = x0 * page_rect.width, x1 * page_rect.width
                    y0, y1 = y0 * page_rect.height, y1 * page_rect.height
                rect = fitz.Rect(x0, y0, x1, y1)
                if rect.width <= 0 or rect.height <= 0:
                    continue
                annotation = page.add_highlight_annot(rect)
                if annotation is not None:
                    annotation.update()
                did_highlight = True
                selected_page_candidate = _select_highlight_page_candidate(
                    selected_page_candidate,
                    field=target.get("field"),
                    target_index=target_index,
                    page=page_number,
                )
                continue

            search_value = target.get("value")
            if not isinstance(search_value, str) or not search_value.strip():
                continue
            page_number = _optional_positive_int(target.get("page"))
            if page_number is not None:
                page_indexes = [page_number - 1]
            else:
                page_indexes = list(range(document.page_count))
            for page_index in page_indexes:
                if page_index < 0 or page_index >= document.page_count:
                    continue
                page = document[page_index]
                matches = page.search_for(search_value.strip())
                for match in matches:
                    annotation = page.add_highlight_annot(match)
                    if annotation is not None:
                        annotation.update()
                if matches:
                    selected_page_candidate = _select_highlight_page_candidate(
                        selected_page_candidate,
                        field=target.get("field"),
                        target_index=target_index,
                        page=page_number or (page_index + 1),
                    )
                    did_highlight = True

        if not did_highlight:
            raise ValueError("Unable to locate requested source content in PDF")
        selected_page = selected_page_candidate[2] if selected_page_candidate else 1
        return document.tobytes(garbage=3, deflate=True), selected_page
    finally:
        document.close()


async def execute(
    *,
    supabase: SupabaseNamespacedPort,
    collabora: CollaboraPort,
    source_file_id: str,
    sheet: str | None = None,
    cell: str | None = None,
    cell_range: str | None = None,
    source_refs: list[dict[str, Any]] | None = None,
    preferred_sheet: str | None = None,
    highlight_file_id: str | None = None,
) -> dict[str, Any]:
    file_entry = supabase.file.get_file(source_file_id)
    if not file_entry:
        raise LookupError("Source file not found")

    content = file_entry.get("content")
    if not isinstance(content, (bytes, bytearray)):
        raise ValueError("Stored source file content is invalid")
    source_bytes = bytes(content)
    source_filename = _optional_str(file_entry, "name") or f"{source_file_id}.xlsx"
    source_content_type = _optional_str(file_entry, "content_type")

    source_format = classify_document(
        filename=source_filename,
        content_type=source_content_type,
        file_bytes=source_bytes,
    )
    if not source_format.is_spreadsheet:
        if not source_refs:
            raise ValueError(
                "Missing non-spreadsheet source references (expected page+bbox, page+value, or value)"
            )
        pdf_targets = _collect_pdf_targets(source_refs)
        if source_format.kind == "pdf":
            source_pdf_bytes = source_bytes
        else:
            source_pdf_bytes = await collabora.convert_document_to_pdf_collabora(
                source_bytes,
                filename=source_filename,
                content_type=source_content_type or "application/octet-stream",
                collabora_base_url=os.getenv("COLLABORA_URL", "http://localhost:8080"),
            )
        highlighted_pdf_bytes, highlighted_page = _highlight_pdf_bytes(
            pdf_bytes=source_pdf_bytes,
            targets=pdf_targets,
        )
        output_file_id = (highlight_file_id or "").strip() or str(uuid.uuid4())
        base_name, _ = os.path.splitext(source_filename)
        output_filename = f"{base_name or 'document'}-source-highlight.pdf"
        supabase.file.save_file(
            file_id=output_file_id,
            name=output_filename,
            content=highlighted_pdf_bytes,
            content_type=_PDF_CONTENT_TYPE,
            file_origin="source_highlight",
        )
        return {
            "file_id": output_file_id,
            "filename": output_filename,
            "page": highlighted_page,
        }

    coordinate_input = (cell_range or cell or "").strip()
    if not coordinate_input and not source_refs:
        raise ValueError("Missing source coordinate (cell or cell_range)")

    source_sheet: str | None = None
    source_ref = ""
    normalized_sheet = _normalize_sheet_name((sheet or "").strip())
    if coordinate_input:
        source_sheet, source_ref = _split_sheet_and_ref(coordinate_input)
        normalized_sheet = _normalize_sheet_name(source_sheet or (sheet or "").strip())

    keep_vba = source_filename.lower().endswith(".xlsm")
    workbook = load_workbook(io.BytesIO(source_bytes), keep_vba=keep_vba)
    try:
        targets: list[dict[str, str]] = []
        if source_refs:
            for source_ref_entry in source_refs:
                if not isinstance(source_ref_entry, dict):
                    continue
                source_field = source_ref_entry.get("field")
                field = (
                    source_field.lower().strip()
                    if isinstance(source_field, str) and source_field.strip()
                    else None
                )
                source_sheet_hint = source_ref_entry.get("sheet")
                sheet_hint = (
                    str(source_sheet_hint)
                    if isinstance(source_sheet_hint, str) and source_sheet_hint.strip()
                    else None
                )
                source_cell_range = source_ref_entry.get("cell_range")
                source_cell = source_ref_entry.get("cell")
                raw_ref_value = (
                    source_cell_range
                    if isinstance(source_cell_range, str) and source_cell_range.strip()
                    else source_cell
                    if isinstance(source_cell, str) and source_cell.strip()
                    else None
                )
                if not raw_ref_value:
                    continue
                target = _build_target(
                    workbook=workbook,
                    raw_sheet=sheet_hint,
                    raw_ref=str(raw_ref_value),
                    fallback_sheet=preferred_sheet or sheet,
                    field=field,
                )
                if target:
                    targets.append(target)

        if not targets:
            target_ref = source_ref or coordinate_input
            fallback_target = _build_target(
                workbook=workbook,
                raw_sheet=normalized_sheet or sheet,
                raw_ref=target_ref,
                fallback_sheet=preferred_sheet,
            )
            if not fallback_target:
                fallback_sheet_name = _normalize_sheet_name(
                    normalized_sheet or preferred_sheet or workbook.active.title
                )
                raise ValueError(
                    f"Sheet '{fallback_sheet_name}' was not found in source workbook"
                )
            targets.append(fallback_target)

        deduped_targets: list[dict[str, str]] = []
        seen_targets: set[tuple[str, str]] = set()
        for target in targets:
            key = (target["sheet"], target["cell_range"])
            if key in seen_targets:
                continue
            seen_targets.add(key)
            deduped_targets.append(target)
        targets = deduped_targets

        for target in targets:
            worksheet = workbook[target["sheet"]]
            normalized_target_range = target["cell_range"]
            min_col, min_row, max_col, max_row = range_boundaries(normalized_target_range)
            for row in worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for current_cell in row:
                    current_cell.fill = _HIGHLIGHT_FILL

        title_target = next((item for item in targets if item.get("field") == "title"), None)
        preferred_sheet_name = _normalize_sheet_name(preferred_sheet or "")
        preferred_target = (
            next((item for item in targets if item["sheet"] == preferred_sheet_name), None)
            if preferred_sheet_name
            else None
        )
        selected_target = title_target or preferred_target or targets[0]
        target_sheet = selected_target["sheet"]
        normalized_target_range = selected_target["cell_range"]
        worksheet = workbook[target_sheet]
        min_col, min_row, _, _ = range_boundaries(normalized_target_range)
        top_left = f"{get_column_letter(min_col)}{min_row}"
        if worksheet.sheet_view.selection:
            worksheet.sheet_view.selection[0].activeCell = top_left
            worksheet.sheet_view.selection[0].sqref = normalized_target_range
        workbook.active = workbook.sheetnames.index(target_sheet)

        output = io.BytesIO()
        workbook.save(output)
    finally:
        workbook.close()

    output_file_id = (highlight_file_id or "").strip() or str(uuid.uuid4())
    base_name, _ = os.path.splitext(source_filename)
    output_ext = ".xlsm" if keep_vba else ".xlsx"
    output_filename = f"{base_name or 'document'}-source-highlight{output_ext}"
    output_content_type = _XLSM_CONTENT_TYPE if keep_vba else _XLSX_CONTENT_TYPE

    supabase.file.save_file(
        file_id=output_file_id,
        name=output_filename,
        content=output.getvalue(),
        content_type=output_content_type,
        file_origin="source_highlight",
    )

    return {
        "file_id": output_file_id,
        "filename": output_filename,
        "sheet": target_sheet,
        "cell_range": normalized_target_range,
    }
