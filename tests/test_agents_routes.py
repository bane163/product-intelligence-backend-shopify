import io
import json
import uuid
from typing import Any

import openpyxl
import pytest
from httpx import AsyncClient, ASGITransport
from PIL import Image

from app_context import get_app_context
from api.agents.files_helper import _generate_thumbnail_bytes, _is_blank_png
from main import app

TEST_SHOP_DOMAIN = "store.myshopify.com"
TEST_SHOP_HEADERS = {"x-shop-domain": TEST_SHOP_DOMAIN}


def _force_in_memory_storage(monkeypatch):
    ctx = get_app_context()
    service = getattr(ctx.services.supabase, "_service", None)
    if service is not None:
        monkeypatch.setattr(service, "_try_get_bucket", lambda: None, raising=False)
        monkeypatch.setattr(service, "_get_supabase_client", lambda: None, raising=False)
    monkeypatch.setattr(
        ctx.services.supabase, "_try_get_bucket", lambda: None, raising=False
    )
    monkeypatch.setattr(
        ctx.services.supabase, "_get_supabase_client", lambda: None, raising=False
    )


@pytest.mark.asyncio
async def test_upload_and_wopi_get_file_contents():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        file_bytes = b"hello world"
        files = {
            "file": (
                "test.xlsx",
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        r = await ac.post("/agents/upload", files=files)
        assert r.status_code == 200
        body = r.json()
        assert "file_id" in body
        file_id = body["file_id"]

        # GET contents via WOPI
        r2 = await ac.get(f"/agents/wopi/files/{file_id}/contents")
        assert r2.status_code == 200
        assert r2.content == file_bytes


@pytest.mark.asyncio
async def test_file_info_and_delete():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        file_bytes = b"abc123"
        files = {
            "file": (
                "doc.xlsx",
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        r = await ac.post("/agents/upload", files=files)
        assert r.status_code == 200
        file_id = r.json()["file_id"]

        r_info = await ac.get(f"/agents/files/{file_id}")
        assert r_info.status_code == 200
        info = r_info.json()
        assert info["filename"] == "doc.xlsx"

        r_del = await ac.delete(f"/agents/files/{file_id}")
        assert r_del.status_code == 200
        assert r_del.json()["status"] == "deleted"

        # now should 404
        r_info2 = await ac.get(f"/agents/files/{file_id}")
        assert r_info2.status_code == 404


@pytest.mark.asyncio
async def test_bulk_delete_files():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        ids: list[str] = []
        for name in ("doc-a.xlsx", "doc-b.xlsx"):
            files = {
                "file": (
                    name,
                    io.BytesIO(b"bulk-delete"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            uploaded = await ac.post("/agents/upload", files=files)
            assert uploaded.status_code == 200
            ids.append(uploaded.json()["file_id"])

        missing_id = "missing-file-id"
        bulk_response = await ac.post(
            "/agents/files/bulk-delete", json={"ids": [ids[0], missing_id, ids[1]]}
        )
        assert bulk_response.status_code == 200
        body = bulk_response.json()
        assert body["deleted_ids"] == [ids[0], missing_id, ids[1]]
        assert body["failed_ids"] == []

        assert (await ac.get(f"/agents/files/{ids[0]}")).status_code == 404
        assert (await ac.get(f"/agents/files/{ids[1]}")).status_code == 404


@pytest.mark.asyncio
async def test_upload_csv_converts_to_xlsx(monkeypatch):
    async def fake_convert_csv_to_excel(content, collabora_base_url=None, timeout=60):
        _ = (content, collabora_base_url, timeout)
        return b"XLSX_BYTES"

    import ai.collabora_utils as cu

    monkeypatch.setattr(cu, "convert_csv_to_excel", fake_convert_csv_to_excel)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {"file": ("products.csv", io.BytesIO(b"a,b\n1,2\n"), "text/csv")}
        r = await ac.post("/agents/upload", files=files)
        assert r.status_code == 200
        file_id = r.json()["file_id"]

        r_info = await ac.get(f"/agents/files/{file_id}")
        assert r_info.status_code == 200
        info = r_info.json()
        assert info["filename"] == "products.xlsx"
        assert (
            info["content_type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert info["size"] == len(b"XLSX_BYTES")


@pytest.mark.asyncio
async def test_upload_rejects_unsupported_file_type():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": ("unsupported.zip", io.BytesIO(b"zip-data"), "application/zip")
        }
        response = await ac.post("/agents/upload", files=files)
        assert response.status_code == 415
        assert "Unsupported file type" in response.json()["detail"]


@pytest.mark.asyncio
async def test_upload_csv_conversion_failure_returns_500_and_skips_save(monkeypatch):
    async def fake_convert_csv_to_excel(content, collabora_base_url=None, timeout=60):
        _ = (content, collabora_base_url, timeout)
        raise RuntimeError("conversion failed")

    import ai.collabora_utils as cu

    monkeypatch.setattr(cu, "convert_csv_to_excel", fake_convert_csv_to_excel)

    ctx = get_app_context()
    save_called = {"value": False}

    def fail_if_save_called(*args, **kwargs):
        _ = (args, kwargs)
        save_called["value"] = True
        raise AssertionError("save_file should not be called when CSV conversion fails")

    monkeypatch.setattr(ctx.services.supabase, "save_file", fail_if_save_called)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {"file": ("products.csv", io.BytesIO(b"a,b\n1,2\n"), "text/csv")}
        r = await ac.post("/agents/upload", files=files)
        assert r.status_code == 500
        assert "CSV conversion failed" in r.json()["detail"]
        assert save_called["value"] is False


@pytest.mark.asyncio
async def test_upload_does_not_block_on_thumbnail_generation(monkeypatch):
    def fail_if_thumbnail_called(*args, **kwargs):
        _ = (args, kwargs)
        raise AssertionError("upload route should not generate thumbnails synchronously")

    import api.agents.files as files_api

    monkeypatch.setattr(files_api, "generate_thumbnail_bytes", fail_if_thumbnail_called)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "sheet.xlsx",
                io.BytesIO(b"sheetdata"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        r = await ac.post("/agents/upload", files=files)
        assert r.status_code == 200
        body = r.json()
        assert "file_id" in body
        assert body["thumbnail_generated"] is False


@pytest.mark.asyncio
async def test_bulk_upload_returns_per_file_results_with_partial_failure():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = [
            (
                "files",
                (
                    "good.xlsx",
                    io.BytesIO(b"good-content"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
            ("files", ("bad.zip", io.BytesIO(b"zip-data"), "application/zip")),
        ]
        response = await ac.post("/agents/upload/bulk", files=files)
        assert response.status_code == 200
        payload = response.json()
        assert payload["total"] == 2
        assert payload["succeeded"] == 1
        assert payload["failed"] == 1
        assert len(payload["uploaded"]) == 1
        assert len(payload["errors"]) == 1
        assert payload["uploaded"][0]["filename"] == "good.xlsx"
        assert payload["errors"][0]["filename"] == "bad.zip"
        assert "Unsupported file type" in payload["errors"][0]["error"]


@pytest.mark.asyncio
async def test_bulk_upload_succeeds_for_multiple_files():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = [
            (
                "files",
                (
                    "one.xlsx",
                    io.BytesIO(b"one-content"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
            (
                "files",
                (
                    "two.xlsx",
                    io.BytesIO(b"two-content"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            ),
        ]
        response = await ac.post("/agents/upload/bulk", files=files)
        assert response.status_code == 200
        payload = response.json()
        assert payload["total"] == 2
        assert payload["succeeded"] == 2
        assert payload["failed"] == 0
        assert len(payload["uploaded"]) == 2
        assert payload["errors"] == []
        uploaded_filenames = [item["filename"] for item in payload["uploaded"]]
        assert uploaded_filenames == ["one.xlsx", "two.xlsx"]


@pytest.mark.asyncio
async def test_preview_generates_png(monkeypatch):
    async def fake_generate_thumbnail_bytes(**kwargs):
        _ = kwargs
        return b"PNG_BYTES_PAGE_1"

    import api.agents.files as files_api

    monkeypatch.setattr(files_api, "generate_thumbnail_bytes", fake_generate_thumbnail_bytes)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        file_bytes = b"sheetdata"
        files = {
            "file": (
                "sheet.xlsx",
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        r = await ac.post("/agents/upload", files=files)
        file_id = r.json()["file_id"]

        r_preview = await ac.get(f"/agents/preview/{file_id}")
        assert r_preview.status_code == 200
        assert r_preview.content == b"PNG_BYTES_PAGE_1"


@pytest.mark.asyncio
async def test_source_highlight_creates_highlighted_copy():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet["A1"] = "Title"
    sheet["B1"] = "Vendor"
    sheet["A2"] = "Demo product"
    sheet["B2"] = "Demo vendor"
    source_buffer = io.BytesIO()
    workbook.save(source_buffer)
    workbook.close()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "source.xlsx",
                io.BytesIO(source_buffer.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={"sheet": "Products", "cell_range": "A2:B2"},
        )
        assert highlighted.status_code == 200
        payload = highlighted.json()
        highlight_file_id = payload["file_id"]
        assert highlight_file_id != source_file_id
        assert payload["sheet"] == "Products"
        assert payload["cell_range"] == "A2:B2"

        highlighted_contents = await ac.get(
            f"/agents/wopi/files/{highlight_file_id}/contents"
        )
        assert highlighted_contents.status_code == 200

        highlighted_workbook = openpyxl.load_workbook(
            io.BytesIO(highlighted_contents.content)
        )
        highlighted_sheet = highlighted_workbook["Products"]
        assert highlighted_sheet["A2"].fill.fill_type == "solid"
        assert highlighted_sheet["B2"].fill.fill_type == "solid"
        highlighted_workbook.close()


@pytest.mark.asyncio
async def test_source_highlight_reuses_existing_highlight_file():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet["A1"] = "Title"
    sheet["A2"] = "First row"
    sheet["A3"] = "Second row"
    source_buffer = io.BytesIO()
    workbook.save(source_buffer)
    workbook.close()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "source.xlsx",
                io.BytesIO(source_buffer.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        first_highlight = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={"sheet": "Products", "cell_range": "A2:A2"},
        )
        assert first_highlight.status_code == 200
        highlight_file_id = first_highlight.json()["file_id"]

        second_highlight = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={
                "sheet": "Products",
                "cell_range": "A3:A3",
                "highlight_file_id": highlight_file_id,
            },
        )
        assert second_highlight.status_code == 200
        assert second_highlight.json()["file_id"] == highlight_file_id

        highlighted_contents = await ac.get(
            f"/agents/wopi/files/{highlight_file_id}/contents"
        )
        assert highlighted_contents.status_code == 200

        highlighted_workbook = openpyxl.load_workbook(
            io.BytesIO(highlighted_contents.content)
        )
        highlighted_sheet = highlighted_workbook["Products"]
        assert highlighted_sheet["A3"].fill.fill_type == "solid"
        assert highlighted_sheet["A2"].fill.fill_type != "solid"
        highlighted_workbook.close()


@pytest.mark.asyncio
async def test_list_files_hides_source_highlight_artifacts():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet["A1"] = "Title"
    sheet["A2"] = "Demo product"
    source_buffer = io.BytesIO()
    workbook.save(source_buffer)
    workbook.close()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "source.xlsx",
                io.BytesIO(source_buffer.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={"sheet": "Products", "cell_range": "A2:A2"},
        )
        assert highlighted.status_code == 200
        highlight_file_id = highlighted.json()["file_id"]

        listed = await ac.get("/agents/files?limit=1000")
        assert listed.status_code == 200
        listed_ids = {str(item.get("file_id")) for item in listed.json().get("files", [])}
        assert source_file_id in listed_ids
        assert highlight_file_id not in listed_ids


@pytest.mark.asyncio
async def test_list_files_hides_draft_resume_files():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "merchant-upload.xlsx",
                    io.BytesIO(b"merchant-upload"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        merchant_file_id = uploaded.json()["file_id"]

        created_draft = await ac.post(
            "/agents/product-drafts",
            data={
                "products_json": json.dumps([{"title": "Resume Product"}]),
                "run_id": "run-resume-hide",
                "import_mode": "auto",
                "draft_name": "resume-hide.xlsx",
            },
        )
        assert created_draft.status_code == 200
        draft_id = created_draft.json()["draft_id"]

        resumed = await ac.post(f"/agents/product-drafts/{draft_id}/resume-file")
        assert resumed.status_code == 200
        resume_file_id = resumed.json()["file_id"]

        listed = await ac.get("/agents/files?limit=1000")
        assert listed.status_code == 200
        listed_ids = {str(item.get("file_id")) for item in listed.json().get("files", [])}
        assert merchant_file_id in listed_ids
        assert resume_file_id not in listed_ids


@pytest.mark.asyncio
async def test_source_highlight_source_refs_highlights_all_cells_and_prefers_title_sheet():
    workbook = openpyxl.Workbook()
    products_sheet = workbook.active
    products_sheet.title = "Products"
    products_sheet["A2"] = "Product one"
    products_sheet["D2"] = "19.99"
    attributes_sheet = workbook.create_sheet("Attributes")
    attributes_sheet["B4"] = "Material"
    attributes_sheet["C4"] = "Cotton"
    source_buffer = io.BytesIO()
    workbook.save(source_buffer)
    workbook.close()

    source_refs = [
        {"field": "title", "sheet": "Products", "cell": "A2"},
        {"field": "price", "sheet": "Products", "cell": "D2"},
        {"field": "material", "sheet": "Attributes", "cell_range": "B4:C4"},
    ]

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "source.xlsx",
                io.BytesIO(source_buffer.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={"source_refs_json": json.dumps(source_refs)},
        )
        assert highlighted.status_code == 200
        payload = highlighted.json()
        assert payload["sheet"] == "Products"
        assert payload["cell_range"] == "A2"

        highlighted_contents = await ac.get(
            f"/agents/wopi/files/{payload['file_id']}/contents"
        )
        highlighted_workbook = openpyxl.load_workbook(
            io.BytesIO(highlighted_contents.content)
        )
        highlighted_products = highlighted_workbook["Products"]
        highlighted_attributes = highlighted_workbook["Attributes"]
        assert highlighted_products["A2"].fill.fill_type == "solid"
        assert highlighted_products["D2"].fill.fill_type == "solid"
        assert highlighted_attributes["B4"].fill.fill_type == "solid"
        assert highlighted_attributes["C4"].fill.fill_type == "solid"
        highlighted_workbook.close()


@pytest.mark.asyncio
async def test_source_highlight_source_refs_prefers_requested_sheet_when_no_title_ref():
    workbook = openpyxl.Workbook()
    details_sheet = workbook.active
    details_sheet.title = "Details"
    details_sheet["A2"] = "SKU-1"
    pricing_sheet = workbook.create_sheet("Pricing")
    pricing_sheet["C3"] = "29.99"
    source_buffer = io.BytesIO()
    workbook.save(source_buffer)
    workbook.close()

    source_refs = [
        {"field": "sku", "sheet": "Details", "cell": "A2"},
        {"field": "price", "sheet": "Pricing", "cell": "C3"},
    ]

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "source.xlsx",
                io.BytesIO(source_buffer.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={
                "source_refs_json": json.dumps(source_refs),
                "preferred_sheet": "Pricing",
            },
        )
        assert highlighted.status_code == 200
        payload = highlighted.json()
        assert payload["sheet"] == "Pricing"
        assert payload["cell_range"] == "C3"

        highlighted_contents = await ac.get(
            f"/agents/wopi/files/{payload['file_id']}/contents"
        )
        highlighted_workbook = openpyxl.load_workbook(
            io.BytesIO(highlighted_contents.content)
        )
        highlighted_details = highlighted_workbook["Details"]
        highlighted_pricing = highlighted_workbook["Pricing"]
        assert highlighted_details["A2"].fill.fill_type == "solid"
        assert highlighted_pricing["C3"].fill.fill_type == "solid"
        highlighted_workbook.close()


@pytest.mark.asyncio
async def test_source_highlight_rejects_invalid_source_refs_json():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet["A1"] = "Title"
    source_buffer = io.BytesIO()
    workbook.save(source_buffer)
    workbook.close()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "source.xlsx",
                io.BytesIO(source_buffer.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={"source_refs_json": "{bad json"},
        )
        assert highlighted.status_code == 422
        assert "source_refs_json" in highlighted.json()["detail"]


def test_select_highlight_page_candidate_prefers_text_over_image():
    import application.use_cases.files.create_source_highlight_file as source_highlight_use_case

    selected = None
    selected = source_highlight_use_case._select_highlight_page_candidate(
        selected,
        field="image_src",
        target_index=2,
        page=2,
    )
    selected = source_highlight_use_case._select_highlight_page_candidate(
        selected,
        field="title",
        target_index=0,
        page=1,
    )
    assert selected is not None
    assert selected[2] == 1


def test_select_highlight_page_candidate_keeps_first_match_for_same_priority():
    import application.use_cases.files.create_source_highlight_file as source_highlight_use_case

    selected = None
    selected = source_highlight_use_case._select_highlight_page_candidate(
        selected,
        field="vendor",
        target_index=0,
        page=1,
    )
    selected = source_highlight_use_case._select_highlight_page_candidate(
        selected,
        field="vendor",
        target_index=3,
        page=4,
    )
    assert selected is not None
    assert selected[2] == 1


@pytest.mark.asyncio
async def test_source_highlight_handles_pdf_source_refs(monkeypatch):
    def fake_highlight_pdf_bytes(*, pdf_bytes, targets):
        _ = (pdf_bytes, targets)
        return b"%PDF-highlighted", 2

    import application.use_cases.files.create_source_highlight_file as source_highlight_use_case

    monkeypatch.setattr(
        source_highlight_use_case, "_highlight_pdf_bytes", fake_highlight_pdf_bytes
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "document.pdf",
                io.BytesIO(b"%PDF-1.4\nfake\n"),
                "application/pdf",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={
                "source_refs_json": json.dumps(
                    [{"page": 2, "value": "Materials", "document_kind": "pdf"}]
                )
            },
        )
        assert highlighted.status_code == 200
        payload = highlighted.json()
        assert payload["page"] == 2
        assert payload["filename"].endswith(".pdf")

        highlighted_contents = await ac.get(
            f"/agents/wopi/files/{payload['file_id']}/contents"
        )
        assert highlighted_contents.status_code == 200
        assert highlighted_contents.content == b"%PDF-highlighted"


@pytest.mark.asyncio
async def test_source_highlight_handles_pdf_source_refs_without_page(monkeypatch):
    captured: dict[str, object] = {}

    def fake_highlight_pdf_bytes(*, pdf_bytes, targets):
        _ = pdf_bytes
        captured["targets"] = targets
        return b"%PDF-highlighted", 3

    import application.use_cases.files.create_source_highlight_file as source_highlight_use_case

    monkeypatch.setattr(
        source_highlight_use_case, "_highlight_pdf_bytes", fake_highlight_pdf_bytes
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "document.pdf",
                io.BytesIO(b"%PDF-1.4\nfake\n"),
                "application/pdf",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={
                "source_refs_json": json.dumps(
                    [
                        {
                            "field": "image_src",
                            "document_kind": "extracted_text",
                            "value": "https://example.com/images/blue-shirt.jpg",
                        }
                    ]
                )
            },
        )
        assert highlighted.status_code == 200
        payload = highlighted.json()
        assert payload["page"] == 3
        targets = captured.get("targets")
        assert isinstance(targets, list)
        assert any(
            isinstance(target, dict)
            and target.get("kind") == "text"
            and target.get("page") is None
            and target.get("value") == "https://example.com/images/blue-shirt.jpg"
            for target in targets
        )


@pytest.mark.asyncio
async def test_source_highlight_converts_non_pdf_document(monkeypatch):
    conversion_called = {"value": False}

    async def fake_convert_document_to_pdf_collabora(
        file_bytes,
        *,
        filename,
        content_type,
        collabora_base_url="http://localhost:8080",
        timeout=60,
    ):
        _ = (file_bytes, filename, content_type, collabora_base_url, timeout)
        conversion_called["value"] = True
        return b"%PDF-converted"

    def fake_highlight_pdf_bytes(*, pdf_bytes, targets):
        _ = targets
        assert pdf_bytes == b"%PDF-converted"
        return b"%PDF-highlighted", 1

    import ai.collabora_utils as cu
    import application.use_cases.files.create_source_highlight_file as source_highlight_use_case

    monkeypatch.setattr(
        cu, "convert_document_to_pdf_collabora", fake_convert_document_to_pdf_collabora
    )
    monkeypatch.setattr(
        source_highlight_use_case, "_highlight_pdf_bytes", fake_highlight_pdf_bytes
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "document.docx",
                io.BytesIO(b"PK\x03\x04fake-docx"),
                "application/octet-stream",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        highlighted = await ac.post(
            f"/agents/files/{source_file_id}/source-highlight",
            data={"source_refs_json": json.dumps([{"page": 1, "value": "Materials"}])},
        )
        assert highlighted.status_code == 200
        payload = highlighted.json()
        assert payload["page"] == 1
        assert payload["filename"].endswith(".pdf")
        assert conversion_called["value"] is True


@pytest.mark.asyncio
async def test_source_target_resolves_non_spreadsheet_target(monkeypatch):
    async def fake_extract_link_targets_collabora(
        file_bytes,
        *,
        filename,
        content_type,
        collabora_base_url="http://localhost:8080",
        timeout=60,
    ):
        _ = (file_bytes, filename, content_type, collabora_base_url, timeout)
        return {
            "Headings": {
                "Materials": "Materials|region",
                "Care": "Care|region",
            },
            "Bookmarks": {"_toc": "_toc"},
        }

    import ai.collabora_utils as cu

    monkeypatch.setattr(
        cu, "extract_link_targets_collabora", fake_extract_link_targets_collabora
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "document.pdf",
                io.BytesIO(b"%PDF-1.4\nfake\n"),
                "application/pdf",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        resolved = await ac.get(
            f"/agents/files/{source_file_id}/source-target",
            params={"value": "Materials & Care", "document_kind": "pdf", "page": 2},
        )
        assert resolved.status_code == 200
        body = resolved.json()
        assert body["target"] == "Materials|region"
        assert body["matched_label"] == "Materials"
        assert body["matched_group"] == "Headings"
        assert body["reason"] == "matched"


@pytest.mark.asyncio
async def test_source_target_skips_spreadsheet_files(monkeypatch):
    called = {"value": False}

    async def fake_extract_link_targets_collabora(*args, **kwargs):
        _ = (args, kwargs)
        called["value"] = True
        return {"Headings": {"Title": "Title|region"}}

    import ai.collabora_utils as cu

    monkeypatch.setattr(
        cu, "extract_link_targets_collabora", fake_extract_link_targets_collabora
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet["A1"] = "Title"
    source_buffer = io.BytesIO()
    workbook.save(source_buffer)
    workbook.close()

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "source.xlsx",
                io.BytesIO(source_buffer.getvalue()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        uploaded = await ac.post("/agents/upload", files=files)
        assert uploaded.status_code == 200
        source_file_id = uploaded.json()["file_id"]

        resolved = await ac.get(
            f"/agents/files/{source_file_id}/source-target",
            params={"value": "Title", "document_kind": "spreadsheet"},
        )
        assert resolved.status_code == 200
        body = resolved.json()
        assert body["target"] is None
        assert body["reason"] == "spreadsheet"
        assert called["value"] is False


@pytest.mark.asyncio
async def test_save_product_draft():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        payload = {
            "products_json": json.dumps([{"title": "Draft Product"}]),
            "run_id": "run-1",
            "import_mode": "create",
            "draft_name": "products.xlsx",
            "input_file_id": "input-file-1",
            "input_filename": "source.xlsx",
        }
        r = await ac.post("/agents/product-drafts", data=payload)
        assert r.status_code == 200
        body = r.json()
        assert body["product_count"] == 1
        assert body["import_mode"] == "create"
        assert body["draft_name"] == "products.xlsx"
        assert body["input_file_id"] == "input-file-1"
        assert body["input_filename"] == "source.xlsx"
        assert "draft_id" in body


def _patch_bulk_submit(monkeypatch, product_results: list[dict] | None = None):
    """Patch all bulk operation methods on ShopifyAdapter for submit tests."""
    from infrastructure.adapters.shopify_adapter import ShopifyAdapter
    import shopify as shopify_module
    from application.use_cases.processing import submit_products as sp_module

    async def fake_create_staged_upload(self):
        return {
            "url": "https://staged.example.com/upload",
            "resourceUrl": "staged://bulk_import.jsonl",
            "parameters": [{"name": "key", "value": "tmp/bulk_import.jsonl"}],
        }

    async def fake_upload_to_staged_url(self, url, parameters, jsonl_data):
        pass

    async def fake_run_bulk_mutation(self, staged_upload_path):
        return {"id": "gid://shopify/BulkOperation/1", "status": "CREATED"}

    results = product_results or []

    async def fake_wait_for_bulk_operation(self, operation_id, *, poll_interval=5.0, timeout=600.0):
        return {
            "id": operation_id,
            "status": "COMPLETED",
            "url": "https://results.example.com/bulk.jsonl",
            "rootObjectCount": len(results),
            "objectCount": len(results),
        }

    async def fake_download_bulk_results(shopify, result_url):
        return results

    monkeypatch.setattr(ShopifyAdapter, "create_staged_upload", fake_create_staged_upload)
    monkeypatch.setattr(ShopifyAdapter, "upload_to_staged_url", fake_upload_to_staged_url)
    monkeypatch.setattr(ShopifyAdapter, "run_bulk_mutation", fake_run_bulk_mutation)
    monkeypatch.setattr(ShopifyAdapter, "wait_for_bulk_operation", fake_wait_for_bulk_operation)
    monkeypatch.setattr(shopify_module.ShopifyClient, "create_staged_upload", fake_create_staged_upload)
    monkeypatch.setattr(shopify_module.ShopifyClient, "upload_to_staged_url", fake_upload_to_staged_url)
    monkeypatch.setattr(shopify_module.ShopifyClient, "run_bulk_mutation", fake_run_bulk_mutation)
    monkeypatch.setattr(shopify_module.ShopifyClient, "wait_for_bulk_operation", fake_wait_for_bulk_operation)
    monkeypatch.setattr(sp_module, "_download_bulk_results", fake_download_bulk_results)


@pytest.mark.asyncio
async def test_submit_products_auto_mode(monkeypatch):
    """Submit should use bulk operation flow (staged upload → bulk mutation → poll)."""
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {"id": "gid://shopify/Product/2", "title": "Demo"},
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        payload = {
            "products_json": json.dumps([{"title": "Demo"}]),
            "import_mode": "auto",
        }
        r = await ac.post("/agents/submit-products", data=payload)
        assert r.status_code == 200
        assert r.json()["success_count"] == 1


@pytest.mark.asyncio
async def test_submit_products_uses_shopify_port_jsonl_builder(monkeypatch):
    from infrastructure.adapters.shopify_adapter import ShopifyAdapter
    import shopify as shopify_module

    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {"id": "gid://shopify/Product/250", "title": "Demo"},
                "userErrors": [],
            }
        }
    }])

    captured: dict[str, Any] = {"called": False, "count": 0}

    def fake_build_product_set_jsonl(products: list[dict[str, Any]]) -> str:
        captured["called"] = True
        captured["count"] = len(products)
        return '{"input":{"title":"Demo"}}'

    def fail_concrete_client_jsonl_builder(products: list[dict[str, Any]]) -> str:
        _ = products
        raise AssertionError("submit should use the ShopifyPort JSONL builder")

    monkeypatch.setattr(
        ShopifyAdapter,
        "build_product_set_jsonl",
        staticmethod(fake_build_product_set_jsonl),
    )
    monkeypatch.setattr(
        shopify_module.ShopifyClient,
        "build_product_set_jsonl",
        staticmethod(fail_concrete_client_jsonl_builder),
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/submit-products",
            data={
                "products_json": json.dumps([{"title": "Demo"}]),
                "import_mode": "auto",
            },
        )
        assert response.status_code == 200
        body = response.json()
        assert body["success_count"] == 1
        assert captured["called"] is True
        assert captured["count"] == 1


@pytest.mark.asyncio
async def test_submit_products_partial_success_creates_submitted_document(monkeypatch):
    _patch_bulk_submit(monkeypatch, [
        {
            "data": {
                "productSet": {
                    "product": {"id": "gid://shopify/Product/300", "title": "Good Product"},
                    "userErrors": [],
                }
            }
        },
        {
            "data": {
                "productSet": {
                    "product": {"id": None, "title": "Bad Product"},
                    "userErrors": [{"field": ["title"], "message": "Title is invalid"}],
                }
            }
        },
    ])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/submit-products",
            data={
                "products_json": json.dumps([{"title": "Good Product"}, {"title": "Bad Product"}]),
                "import_mode": "auto",
                "document_name": "partial-submit.xlsx",
            },
        )
        assert response.status_code == 200
        body = response.json()
        assert body["success_count"] == 1
        assert body["submitted_id"]
        assert any(item.get("status") == "failed" for item in body["results"])

        submitted_detail = await ac.get(
            f"/agents/submitted-documents/{body['submitted_id']}"
        )
        assert submitted_detail.status_code == 200
        stored_products = submitted_detail.json()["submitted_document"]["products"]
        assert len(stored_products) == 1
        assert stored_products[0]["title"] == "Good Product"


@pytest.mark.asyncio
async def test_submit_products_uses_request_scoped_shop_context(monkeypatch):
    from infrastructure.adapters.shopify_adapter import ShopifyAdapter
    import shopify as shopify_module
    from application.use_cases.processing import submit_products as sp_module

    captured: dict[str, Any] = {}

    async def fail_adapter_create_staged_upload(self):
        _ = self
        raise RuntimeError("adapter client should not be used for request-scoped submit")

    async def fake_create_staged_upload(self):
        captured["shop"] = getattr(self, "shop", None)
        captured["token"] = getattr(self, "_token", None)
        return {
            "url": "https://staged.example.com/upload",
            "resourceUrl": "staged://bulk_import.jsonl",
            "parameters": [{"name": "key", "value": "tmp/bulk_import.jsonl"}],
        }

    async def fake_upload_to_staged_url(self, url, parameters, jsonl_data):
        _ = (self, url, parameters, jsonl_data)

    async def fake_run_bulk_mutation(self, staged_upload_path):
        _ = (self, staged_upload_path)
        return {"id": "gid://shopify/BulkOperation/1", "status": "CREATED"}

    async def fake_wait_for_bulk_operation(self, operation_id, *, poll_interval=5.0, timeout=600.0):
        _ = (self, poll_interval, timeout)
        return {
            "id": operation_id,
            "status": "COMPLETED",
            "url": "https://results.example.com/bulk.jsonl",
            "rootObjectCount": 1,
            "objectCount": 1,
        }

    async def fake_download_bulk_results(shopify, result_url):
        _ = (shopify, result_url)
        return [{
            "data": {
                "productSet": {
                    "product": {"id": "gid://shopify/Product/200", "title": "Scoped"},
                    "userErrors": [],
                }
            }
        }]

    monkeypatch.setattr(ShopifyAdapter, "create_staged_upload", fail_adapter_create_staged_upload)
    monkeypatch.setattr(shopify_module.ShopifyClient, "create_staged_upload", fake_create_staged_upload)
    monkeypatch.setattr(shopify_module.ShopifyClient, "upload_to_staged_url", fake_upload_to_staged_url)
    monkeypatch.setattr(shopify_module.ShopifyClient, "run_bulk_mutation", fake_run_bulk_mutation)
    monkeypatch.setattr(shopify_module.ShopifyClient, "wait_for_bulk_operation", fake_wait_for_bulk_operation)
    monkeypatch.setattr(sp_module, "_download_bulk_results", fake_download_bulk_results)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/submit-products",
            data={
                "products_json": json.dumps([{"title": "Scoped"}]),
                "import_mode": "auto",
                "shop_domain": "scoped-shop.myshopify.com",
                "shop_access_token": "shpat_scoped_token",
            },
        )
        assert response.status_code == 200
        body = response.json()
        assert body["success_count"] == 1
        assert captured["shop"] == "scoped-shop.myshopify.com"
        assert captured["token"] == "shpat_scoped_token"


@pytest.mark.asyncio
async def test_submit_products_infers_shop_from_draft_when_request_shop_missing(monkeypatch):
    from infrastructure.adapters.shopify_adapter import ShopifyAdapter
    import shopify as shopify_module
    from application.use_cases.processing import submit_products as sp_module

    captured: dict[str, Any] = {}

    async def fail_adapter_create_staged_upload(self):
        _ = self
        raise RuntimeError("adapter client should not be used when draft has tenant")

    async def fake_create_staged_upload(self):
        captured["shop"] = getattr(self, "shop", None)
        return {
            "url": "https://staged.example.com/upload",
            "resourceUrl": "staged://bulk_import.jsonl",
            "parameters": [{"name": "key", "value": "tmp/bulk_import.jsonl"}],
        }

    async def fake_upload_to_staged_url(self, url, parameters, jsonl_data):
        _ = (self, url, parameters, jsonl_data)

    async def fake_run_bulk_mutation(self, staged_upload_path):
        _ = (self, staged_upload_path)
        return {"id": "gid://shopify/BulkOperation/1", "status": "CREATED"}

    async def fake_wait_for_bulk_operation(self, operation_id, *, poll_interval=5.0, timeout=600.0):
        _ = (self, poll_interval, timeout)
        return {
            "id": operation_id,
            "status": "COMPLETED",
            "url": "https://results.example.com/bulk.jsonl",
            "rootObjectCount": 1,
            "objectCount": 1,
        }

    async def fake_download_bulk_results(shopify, result_url):
        _ = (shopify, result_url)
        return [{
            "data": {
                "productSet": {
                    "product": {"id": "gid://shopify/Product/201", "title": "Draft Scoped"},
                    "userErrors": [],
                }
            }
        }]

    monkeypatch.setattr(ShopifyAdapter, "create_staged_upload", fail_adapter_create_staged_upload)
    monkeypatch.setattr(shopify_module.ShopifyClient, "create_staged_upload", fake_create_staged_upload)
    monkeypatch.setattr(shopify_module.ShopifyClient, "upload_to_staged_url", fake_upload_to_staged_url)
    monkeypatch.setattr(shopify_module.ShopifyClient, "run_bulk_mutation", fake_run_bulk_mutation)
    monkeypatch.setattr(shopify_module.ShopifyClient, "wait_for_bulk_operation", fake_wait_for_bulk_operation)
    monkeypatch.setattr(sp_module, "_download_bulk_results", fake_download_bulk_results)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        created_draft = await ac.post(
            "/agents/product-drafts",
            data={
                "products_json": json.dumps([{"title": "Draft Scoped"}]),
                "run_id": "run-draft-scoped",
                "import_mode": "auto",
                "shop_domain": "draft-scoped.myshopify.com",
                "draft_name": "draft-scoped.xlsx",
            },
        )
        assert created_draft.status_code == 200
        draft_id = created_draft.json()["draft_id"]

        submitted = await ac.post(
            "/agents/submit-products",
            data={
                "draft_id": draft_id,
                "import_mode": "auto",
            },
        )
        assert submitted.status_code == 200
        assert submitted.json()["success_count"] == 1
        assert captured["shop"] == "draft-scoped.myshopify.com"


@pytest.mark.asyncio
async def test_submit_products_auto_updates_when_id_present(monkeypatch):
    """Submit with a product that has an ID should use productSet with identifier."""
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {
                    "id": "gid://shopify/Product/99",
                    "title": "Existing Demo",
                },
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        payload = {
            "products_json": json.dumps(
                [{"id": "gid://shopify/Product/99", "title": "Existing Demo"}]
            ),
            "import_mode": "auto",
        }
        r = await ac.post("/agents/submit-products", data=payload)
        assert r.status_code == 200
        body = r.json()
        assert body["success_count"] == 1
        assert body["results"][0]["mode"] == "productSet"


@pytest.mark.asyncio
async def test_submit_products_uses_submitted_document_when_submitted_id_provided(
    monkeypatch,
):
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {
                    "id": "gid://shopify/Product/404",
                    "title": "Stored Demo",
                },
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        seed_response = await ac.post(
            "/agents/submit-products",
            data={
                "products_json": json.dumps([{"title": "Stored Demo"}]),
                "import_mode": "auto",
                "document_name": "stored-demo.xlsx",
            },
        )
        assert seed_response.status_code == 200
        submitted_id = seed_response.json()["submitted_id"]
        assert submitted_id

        submit_response = await ac.post(
            "/agents/submit-products",
            data={
                "submitted_id": submitted_id,
                "import_mode": "auto",
            },
        )
        assert submit_response.status_code == 200
        body = submit_response.json()
        assert body["success_count"] == 1


@pytest.mark.asyncio
async def test_submit_products_uses_draft_when_draft_id_provided(monkeypatch):
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {
                    "id": "gid://shopify/Product/405",
                    "title": "Draft Source Demo",
                },
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        draft_response = await ac.post(
            "/agents/product-drafts",
            data={
                "products_json": json.dumps([{"title": "Draft Source Demo"}]),
                "run_id": "run-draft-source",
                "import_mode": "auto",
                "draft_name": "draft-source.xlsx",
            },
        )
        assert draft_response.status_code == 200
        draft_id = draft_response.json()["draft_id"]
        assert draft_id

        submit_response = await ac.post(
            "/agents/submit-products",
            data={
                "draft_id": draft_id,
                "import_mode": "auto",
            },
        )
        assert submit_response.status_code == 200
        body = submit_response.json()
        assert body["success_count"] == 1


@pytest.mark.asyncio
async def test_submit_products_rejects_unknown_draft_id(monkeypatch):
    _patch_bulk_submit(monkeypatch, [])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/submit-products",
            data={
                "draft_id": "missing-draft-id",
                "import_mode": "auto",
            },
        )
        assert response.status_code == 400
        assert "Draft not found" in response.json()["detail"]


@pytest.mark.asyncio
async def test_submit_products_rejects_unknown_submitted_id(monkeypatch):
    _patch_bulk_submit(monkeypatch, [])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/submit-products",
            data={
                "submitted_id": "missing-submitted-id",
                "import_mode": "auto",
            },
        )
        assert response.status_code == 400
        assert "Submitted document not found" in response.json()["detail"]


@pytest.mark.asyncio
async def test_submit_products_requires_products_source(monkeypatch):
    async def fail_create_product_from_input(self, product):
        _ = (self, product)
        raise AssertionError(
            "create_product_from_input should not be called without products source"
        )

    import api.agents.submit as submit_api

    monkeypatch.setattr(
        submit_api.ShopifyClient,
        "create_product_from_input",
        fail_create_product_from_input,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/submit-products",
            data={"import_mode": "auto"},
        )
        assert response.status_code == 400
        assert "No products provided" in response.json()["detail"]


@pytest.mark.asyncio
async def test_submit_products_ignores_legacy_ai_enhancements_flag(monkeypatch):
    async def fail_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = supabase, products, shop_domain, normalization_settings, trace_event
        raise AssertionError("submit should not call product intelligence generation")

    import application.use_cases.processing.submit_products as submit_uc

    monkeypatch.setattr(
        submit_uc,
        "generate_suggestions_execute",
        fail_generate_suggestions_execute,
        raising=False,
    )
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {"id": "gid://shopify/Product/555", "title": "Demo"},
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        payload = {
            "products_json": json.dumps([{"title": "Demo"}]),
            "import_mode": "auto",
            "enable_ai_enhancements": "true",
        }
        response = await ac.post("/agents/submit-products", data=payload)
        assert response.status_code == 200
        body = response.json()
        assert body["success_count"] == 1


@pytest.mark.asyncio
async def test_submit_products_legacy_ai_flag_does_not_mutate_draft(monkeypatch):
    async def fake_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = (supabase, products, shop_domain, normalization_settings, trace_event)
        return [
            {
                "product_index": 0,
                "patch_payload": {
                    "vendor": "Synced Vendor",
                    "tags": "synced-tag",
                    "seo_title": "Synced SEO",
                },
            }
        ]

    import application.use_cases.processing.submit_products as submit_uc

    monkeypatch.setattr(
        submit_uc,
        "generate_suggestions_execute",
        fake_generate_suggestions_execute,
        raising=False,
    )
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {"id": "gid://shopify/Product/777", "title": "Drafted Product"},
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        created_draft = await ac.post(
            "/agents/product-drafts",
                data={
                    "products_json": json.dumps([{"title": "Drafted Product"}]),
                    "run_id": "run-sync-draft",
                    "import_mode": "auto",
                    "shop_domain": "store.myshopify.com",
                    "draft_name": "draft-sync.xlsx",
                    "output_file_id": "old-output-file",
                    "output_filename": "old-output.xlsx",
                },
        )
        assert created_draft.status_code == 200
        draft_id = created_draft.json()["draft_id"]

        submitted = await ac.post(
            "/agents/submit-products",
            data={
                "products_json": json.dumps([{"title": "Drafted Product"}]),
                "run_id": "run-sync-submit",
                "import_mode": "auto",
                "draft_id": draft_id,
                "document_name": "draft-sync.xlsx",
                "shop_domain": "store.myshopify.com",
                "enable_ai_enhancements": "true",
            },
        )
        assert submitted.status_code == 200

        draft_detail = await ac.get(f"/agents/product-drafts/{draft_id}")
        assert draft_detail.status_code == 200
        draft = draft_detail.json()["draft"]
        assert draft["products"][0]["title"] == "Drafted Product"
        assert "vendor" not in draft["products"][0]
        assert "seo_title" not in draft["products"][0]
        assert "tags" not in draft["products"][0]
        assert draft.get("output_file_id") == "old-output-file"
        assert draft.get("output_filename") == "old-output.xlsx"


@pytest.mark.asyncio
async def test_import_ignores_freeform_prompt_fields(monkeypatch):
    captured: dict[str, Any] = {}

    async def fake_process_document_execute(**kwargs):
        captured.update(kwargs)
        return {"run_id": "run-import", "result": {"status": "ok"}}

    import application.use_cases.processing.process_document as process_document_uc

    monkeypatch.setattr(process_document_uc, "execute", fake_process_document_execute)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        files = {
            "file": (
                "import.xlsx",
                io.BytesIO(b"sheet"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        response = await ac.post(
            "/agents/import",
            data={"prompt": "Ignore this", "writer_prompt": "Ignore this too"},
            files=files,
        )
        assert response.status_code == 200
        assert "prompt" not in captured
        assert "writer_prompt" not in captured
        assert captured.get("extraction_mode") == "per_sheet"


@pytest.mark.asyncio
async def test_batch_extract_submit_contract_returns_expected_shape_with_partial_failure(
    monkeypatch,
):
    _force_in_memory_storage(monkeypatch)
    tenant_header = {"x-shop-domain": "batch-contract.myshopify.com"}
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-contract.xlsx",
                    io.BytesIO(b"batch-contract"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        valid_file_id = uploaded.json()["file_id"]

        response = await ac.post(
            "/agents/import/batch",
            headers=tenant_header,
            json={
                "file_ids": [valid_file_id, "missing-file-id"],
                "import_mode": "auto",
                "extraction_mode": "per_sheet",
                "auto_submit": True,
                "offload": True,
            },
        )
        assert response.status_code == 202
        body = response.json()
        assert body["total"] == 2
        assert body["queued"] == 1
        assert body["failed"] == 1
        assert len(body["accepted"]) == 1
        assert len(body["errors"]) == 1

        accepted = body["accepted"][0]
        assert accepted["index"] == 0
        assert accepted["file_id"] == valid_file_id
        assert accepted["status"] == "queued"
        assert isinstance(accepted["draft_id"], str) and accepted["draft_id"]
        assert (
            isinstance(accepted["extraction_run_id"], str)
            and accepted["extraction_run_id"]
        )
        assert isinstance(accepted["submit_run_id"], str) and accepted["submit_run_id"]

        error = body["errors"][0]
        assert error["index"] == 1
        assert error["file_id"] == "missing-file-id"
        assert error["code"] == "file_not_found"
        assert "File not found" in error["error"]


@pytest.mark.asyncio
async def test_batch_extract_submit_contract_requires_non_empty_file_ids(monkeypatch):
    _force_in_memory_storage(monkeypatch)
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/import/batch",
            json={"file_ids": [], "auto_submit": True, "offload": True},
        )
        assert response.status_code == 422


@pytest.mark.asyncio
async def test_batch_extract_submit_requires_shop_domain(monkeypatch):
    _force_in_memory_storage(monkeypatch)
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-missing-tenant.xlsx",
                    io.BytesIO(b"batch-missing-tenant"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]

        response = await ac.post(
            "/agents/import/batch",
            json={
                "file_ids": [file_id],
                "import_mode": "auto",
                "extraction_mode": "per_sheet",
                "auto_submit": True,
                "offload": True,
            },
        )
        assert response.status_code == 400
        assert "shop_domain" in response.text


@pytest.mark.asyncio
async def test_batch_extract_submit_persists_tenant_scoped_draft_visible_in_list(
    monkeypatch,
):
    _force_in_memory_storage(monkeypatch)
    tenant_header = {"x-shop-domain": "batch-shop.myshopify.com"}
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-visible.xlsx",
                    io.BytesIO(b"batch-visible"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]

        queued = await ac.post(
            "/agents/import/batch",
            headers=tenant_header,
            json={
                "file_ids": [file_id],
                "import_mode": "auto",
                "extraction_mode": "per_sheet",
                "auto_submit": True,
                "offload": True,
            },
        )
        assert queued.status_code == 202
        draft_id = queued.json()["accepted"][0]["draft_id"]

        listed = await ac.get("/agents/product-drafts", headers=tenant_header)
        assert listed.status_code == 200
        draft_ids = [item.get("draft_id") for item in listed.json()["drafts"]]
        assert draft_id in draft_ids


@pytest.mark.asyncio
async def test_batch_extract_submit_queues_document_import_jobs(monkeypatch):
    _force_in_memory_storage(monkeypatch)
    tenant_header = {"x-shop-domain": "batch-queue.myshopify.com"}
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        assert require_persistent_queue is True
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        first_upload = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-queue-1.xlsx",
                    io.BytesIO(b"batch-queue-1"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        second_upload = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-queue-2.xlsx",
                    io.BytesIO(b"batch-queue-2"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert first_upload.status_code == 200
        assert second_upload.status_code == 200
        first_file_id = first_upload.json()["file_id"]
        second_file_id = second_upload.json()["file_id"]

        response = await ac.post(
            "/agents/import/batch",
            headers=tenant_header,
            json={
                "file_ids": [first_file_id, second_file_id],
                "import_mode": "auto",
                "extraction_mode": "per_sheet",
                "auto_submit": True,
                "offload": True,
            },
        )
        assert response.status_code == 202
        body = response.json()
        assert body["total"] == 2
        assert body["queued"] == 2
        assert body["failed"] == 0
        assert len(body["accepted"]) == 2
        assert len(queued_jobs) == 2

        queued_by_file_id = {item["file_id"]: item for item in queued_jobs}
        assert set(queued_by_file_id.keys()) == {first_file_id, second_file_id}
        for accepted in body["accepted"]:
            job = queued_by_file_id[accepted["file_id"]]
            assert job["job_type"] == "document_import"
            assert job["status"] == "queued"
            assert job["run_id"] == accepted["extraction_run_id"]
            assert job["draft_id"] == accepted["draft_id"]
            assert job["payload"]["auto_submit"] is True
            assert job["payload"]["submit_run_id"] == accepted["submit_run_id"]
            assert job["payload"]["extraction_mode"] == "per_sheet"


@pytest.mark.asyncio
async def test_batch_extract_submit_reuses_active_draft_without_duplicate_enqueue(
    monkeypatch,
):
    _force_in_memory_storage(monkeypatch)
    tenant_header = {"x-shop-domain": "batch-idempotency.myshopify.com"}
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        assert require_persistent_queue is True
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        upload = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-idempotency.xlsx",
                    io.BytesIO(b"batch-idempotency"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert upload.status_code == 200
        file_id = upload.json()["file_id"]

        first = await ac.post(
            "/agents/import/batch",
            headers=tenant_header,
            json={"file_ids": [file_id], "auto_submit": True, "offload": True},
        )
        second = await ac.post(
            "/agents/import/batch",
            headers=tenant_header,
            json={"file_ids": [file_id], "auto_submit": True, "offload": True},
        )
        assert first.status_code == 202
        assert second.status_code == 202
        first_accepted = first.json()["accepted"][0]
        second_accepted = second.json()["accepted"][0]

        assert second_accepted["draft_id"] == first_accepted["draft_id"]
        assert second_accepted["extraction_run_id"] == first_accepted["extraction_run_id"]
        assert second_accepted["submit_run_id"] == first_accepted["submit_run_id"]
        assert len(queued_jobs) == 1


@pytest.mark.asyncio
async def test_batch_extract_submit_reports_per_file_queue_failure(monkeypatch):
    _force_in_memory_storage(monkeypatch)
    tenant_header = {"x-shop-domain": "batch-errors.myshopify.com"}
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []
    failing_file_id = {"value": None}

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        assert require_persistent_queue is True
        payload = {"job_id": job_id, **fields}
        if payload.get("file_id") == failing_file_id["value"]:
            raise RuntimeError("offload_jobs table missing")
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        first_upload = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-queue-ok.xlsx",
                    io.BytesIO(b"batch-queue-ok"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        second_upload = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "batch-queue-fail.xlsx",
                    io.BytesIO(b"batch-queue-fail"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert first_upload.status_code == 200
        assert second_upload.status_code == 200
        first_file_id = first_upload.json()["file_id"]
        failing_file_id["value"] = second_upload.json()["file_id"]

        response = await ac.post(
            "/agents/import/batch",
            headers=tenant_header,
            json={
                "file_ids": [first_file_id, failing_file_id["value"]],
                "auto_submit": True,
                "offload": True,
            },
        )
        assert response.status_code == 202
        body = response.json()
        assert body["total"] == 2
        assert body["queued"] == 1
        assert body["failed"] == 1
        assert len(body["accepted"]) == 1
        assert len(body["errors"]) == 1
        assert body["accepted"][0]["file_id"] == first_file_id
        assert body["errors"][0]["file_id"] == failing_file_id["value"]
        assert body["errors"][0]["code"] == "queue_failed"
        assert "offload_jobs table missing" in body["errors"][0]["error"]
        assert len(queued_jobs) == 1
        assert queued_jobs[0]["file_id"] == first_file_id


@pytest.mark.asyncio
async def test_import_offload_returns_queued_with_draft_tracking(monkeypatch):
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []
    process_calls = 0

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        payload = dict(kwargs)
        payload.pop("require_lifecycle_columns", None)
        return original_save_product_draft(**payload)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        _ = require_persistent_queue
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    async def fake_process_document_execute(**kwargs):
        nonlocal process_calls
        process_calls += 1
        _ = kwargs
        return {
            "run_id": "run-import-offload",
            "result": {"products": [{"title": "Queued Product"}]},
        }

    import application.use_cases.processing.process_document as process_document_uc

    monkeypatch.setattr(process_document_uc, "execute", fake_process_document_execute)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "offload.xlsx",
                    io.BytesIO(b"offload"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]

        response = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-offload", "offload": "true"},
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["run_id"] == "run-import-offload"
        assert body["file_id"] == file_id
        assert isinstance(body.get("draft_id"), str) and body["draft_id"]
        assert len(queued_jobs) == 1
        assert queued_jobs[0]["job_type"] == "document_import"
        assert queued_jobs[0]["run_id"] == "run-import-offload"
        assert queued_jobs[0]["draft_id"] == body["draft_id"]
        assert queued_jobs[0]["file_id"] == file_id
        assert queued_jobs[0]["status"] == "queued"
        assert queued_jobs[0]["payload"]["extraction_mode"] == "per_sheet"
        assert (
            queued_jobs[0]["payload"]["input_content_type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert process_calls == 0

        detail = await ac.get(f"/agents/product-drafts/{body['draft_id']}")
        assert detail.status_code == 200
        draft = detail.json()["draft"]
        assert draft["draft_name"] == "offload.xlsx"
        assert draft["extraction_run_id"] == "run-import-offload"
        assert draft["extraction_status"] == "queued"


@pytest.mark.asyncio
async def test_import_offload_reuses_active_draft_for_same_file(monkeypatch):
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        payload = dict(kwargs)
        payload.pop("require_lifecycle_columns", None)
        return original_save_product_draft(**payload)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        _ = require_persistent_queue
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "offload-reuse.xlsx",
                    io.BytesIO(b"offload-reuse"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]

        first = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-reuse-1", "offload": "true"},
        )
        assert first.status_code == 202
        first_body = first.json()

        second = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-reuse-2", "offload": "true"},
        )
        assert second.status_code == 202
        second_body = second.json()

        assert second_body["draft_id"] == first_body["draft_id"]
        assert second_body["run_id"] == first_body["run_id"]
        assert second_body["status"] == "queued"
        assert len(queued_jobs) == 1


@pytest.mark.asyncio
async def test_import_offload_reuse_backfills_missing_draft_name(monkeypatch):
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        payload = dict(kwargs)
        payload.pop("require_lifecycle_columns", None)
        return original_save_product_draft(**payload)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        _ = require_persistent_queue
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "offload-reuse-naming.xlsx",
                    io.BytesIO(b"offload-reuse-naming"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]

        first = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-reuse-name-1", "offload": "true"},
        )
        assert first.status_code == 202
        first_body = first.json()

        original_save_product_draft(
            draft_id=first_body["draft_id"],
            run_id=first_body["run_id"],
            import_mode="auto",
            draft_name="",
            input_file_id=file_id,
            input_filename="offload-reuse-naming.xlsx",
            extraction_status="queued",
            extraction_run_id=first_body["run_id"],
            extraction_error=None,
            submit_status=None,
            submit_run_id=None,
            submit_error=None,
            products=[],
        )

        second = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-reuse-name-2", "offload": "true"},
        )
        assert second.status_code == 202
        second_body = second.json()

        assert second_body["draft_id"] == first_body["draft_id"]
        assert second_body["run_id"] == first_body["run_id"]
        assert len(queued_jobs) == 1

        detail = await ac.get(f"/agents/product-drafts/{first_body['draft_id']}")
        assert detail.status_code == 200
        draft = detail.json()["draft"]
        assert draft["draft_name"] == "offload-reuse-naming.xlsx"
        assert draft["draft_name"].strip()
        assert draft["draft_name"].lower() != "untitled"


@pytest.mark.asyncio
async def test_import_offload_uses_deterministic_name_when_source_filename_missing(
    monkeypatch,
):
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        payload = dict(kwargs)
        payload.pop("require_lifecycle_columns", None)
        return original_save_product_draft(**payload)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        _ = require_persistent_queue
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    file_id = f"blank-name-{uuid.uuid4()}"
    ctx.services.supabase.save_file(
        file_id=file_id,
        name="",
        content=b"offload",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-no-name", "offload": "true"},
        )
        assert response.status_code == 202
        body = response.json()

        detail = await ac.get(f"/agents/product-drafts/{body['draft_id']}")
        assert detail.status_code == 200
        draft = detail.json()["draft"]
        assert draft["draft_name"] == f"draft-{body['draft_id'][:8]}.xlsx"
        assert draft["draft_name"].strip()
        assert draft["draft_name"].lower() != "untitled"
        assert len(queued_jobs) == 1


@pytest.mark.asyncio
async def test_submit_offload_returns_queued_with_draft_tracking(monkeypatch):
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []
    submit_calls = 0

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        payload = dict(kwargs)
        payload.pop("require_lifecycle_columns", None)
        return original_save_product_draft(**payload)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        _ = require_persistent_queue
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    async def fake_submit_execute(**kwargs):
        nonlocal submit_calls
        submit_calls += 1
        _ = kwargs
        return {"submitted_id": "submitted-offload"}

    import application.use_cases.processing.submit_products as submit_uc

    monkeypatch.setattr(submit_uc, "execute", fake_submit_execute)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        created = await ac.post(
            "/agents/product-drafts",
            data={
                "products_json": json.dumps([{"title": "Queued Draft Product"}]),
                "run_id": "run-submit-base",
                "import_mode": "auto",
                "draft_name": "Queued Draft",
            },
        )
        assert created.status_code == 200
        draft_id = created.json()["draft_id"]

        response = await ac.post(
            "/agents/submit-products",
            data={
                "draft_id": draft_id,
                "run_id": "run-submit-offload",
                "offload": "true",
                "import_mode": "auto",
                "products_json": json.dumps([{"title": "Inline Product"}]),
                "shop_access_token": "shpat_test_token",
                "enable_ai_enhancements": "true",
                "document_name": "queued-submit.xlsx",
            },
        )
        assert response.status_code == 202
        body = response.json()
        assert body["status"] == "queued"
        assert body["draft_id"] == draft_id
        assert body["run_id"] == "run-submit-offload"
        assert len(queued_jobs) == 1
        assert queued_jobs[0]["job_type"] == "shopify_submit"
        assert queued_jobs[0]["run_id"] == "run-submit-offload"
        assert queued_jobs[0]["draft_id"] == draft_id
        assert queued_jobs[0]["status"] == "queued"
        assert queued_jobs[0]["payload"]["products_json"] == json.dumps(
            [{"title": "Inline Product"}]
        )
        assert queued_jobs[0]["payload"]["shop_access_token"] == "shpat_test_token"
        assert "enable_ai_enhancements" not in queued_jobs[0]["payload"]
        assert submit_calls == 0

        detail = await ac.get(f"/agents/product-drafts/{draft_id}")
        assert detail.status_code == 200
        draft = detail.json()["draft"]
        assert draft["submit_run_id"] == "run-submit-offload"
        assert draft["submit_status"] == "queued"


@pytest.mark.asyncio
async def test_submit_offload_resolves_shop_context_from_headers(monkeypatch):
    ctx = get_app_context()
    queued_jobs: list[dict[str, Any]] = []

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        payload = dict(kwargs)
        payload.pop("require_lifecycle_columns", None)
        return original_save_product_draft(**payload)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        _ = require_persistent_queue
        payload = {"job_id": job_id, **fields}
        queued_jobs.append(payload)
        return payload

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    transport = ASGITransport(app=app)
    headers = {
        "x-shop-domain": "header-shop.myshopify.com",
        "x-shop-access-token": "shpat_header_token",
    }
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        created = await ac.post(
            "/agents/product-drafts",
            data={
                "products_json": json.dumps([{"title": "Header Draft Product"}]),
                "run_id": "run-submit-header-base",
                "import_mode": "auto",
                "draft_name": "Header Draft",
            },
            headers=headers,
        )
        assert created.status_code == 200
        draft_id = created.json()["draft_id"]

        response = await ac.post(
            "/agents/submit-products",
            data={
                "draft_id": draft_id,
                "run_id": "run-submit-header-offload",
                "offload": "true",
                "import_mode": "auto",
            },
            headers=headers,
        )
        assert response.status_code == 202
        assert len(queued_jobs) == 1
        assert queued_jobs[0]["shop_domain"] == "header-shop.myshopify.com"
        assert (
            queued_jobs[0]["payload"]["shop_access_token"] == "shpat_header_token"
        )


@pytest.mark.asyncio
async def test_import_offload_returns_explicit_error_when_queue_persistence_fails(
    monkeypatch,
):
    ctx = get_app_context()

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        payload = dict(kwargs)
        payload.pop("require_lifecycle_columns", None)
        return original_save_product_draft(**payload)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    def fake_enqueue_offload_job(
        job_id: str,
        fields: dict[str, Any],
        *,
        require_persistent_queue: bool = False,
    ):
        _ = (job_id, fields)
        assert require_persistent_queue is True
        raise RuntimeError("offload_jobs table missing")

    monkeypatch.setattr(
        ctx.services.supabase,
        "enqueue_offload_job",
        fake_enqueue_offload_job,
        raising=False,
    )

    transport = ASGITransport(app=app, raise_app_exceptions=False)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "offload-failfast.xlsx",
                    io.BytesIO(b"offload"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]

        response = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-failfast", "offload": "true"},
        )
        assert response.status_code == 500
        assert "offload_jobs table missing" in response.json()["detail"]


@pytest.mark.asyncio
async def test_import_offload_returns_explicit_error_when_lifecycle_persistence_fails(
    monkeypatch,
):
    ctx = get_app_context()

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        if kwargs.get("require_lifecycle_columns"):
            raise RuntimeError("product_drafts lifecycle columns missing")
        return original_save_product_draft(**kwargs)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    transport = ASGITransport(app=app, raise_app_exceptions=False)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        uploaded = await ac.post(
            "/agents/upload",
            files={
                "file": (
                    "offload-lifecycle-failfast.xlsx",
                    io.BytesIO(b"offload"),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert uploaded.status_code == 200
        file_id = uploaded.json()["file_id"]

        response = await ac.post(
            "/agents/import",
            data={"file_id": file_id, "run_id": "run-import-lifecycle-failfast", "offload": "true"},
        )
        assert response.status_code == 500
        assert "lifecycle columns missing" in response.json()["detail"]


@pytest.mark.asyncio
async def test_submit_offload_returns_explicit_error_when_lifecycle_persistence_fails(
    monkeypatch,
):
    ctx = get_app_context()
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        created = await ac.post(
            "/agents/product-drafts",
            data={
                "products_json": json.dumps([{"title": "Failfast Draft"}]),
                "run_id": "run-submit-failfast-base",
                "import_mode": "auto",
                "draft_name": "Failfast Draft",
            },
        )
        assert created.status_code == 200
        draft_id = created.json()["draft_id"]

    original_save_product_draft = ctx.services.supabase.save_product_draft

    def fake_save_product_draft(**kwargs):
        if kwargs.get("require_lifecycle_columns"):
            raise RuntimeError("product_drafts lifecycle columns missing")
        return original_save_product_draft(**kwargs)

    monkeypatch.setattr(
        ctx.services.supabase,
        "save_product_draft",
        fake_save_product_draft,
        raising=False,
    )

    transport = ASGITransport(app=app, raise_app_exceptions=False)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.post(
            "/agents/submit-products",
            data={
                "draft_id": draft_id,
                "run_id": "run-submit-failfast",
                "offload": "true",
                "import_mode": "auto",
                "document_name": "failfast-submit.xlsx",
            },
        )
        assert response.status_code == 500
        assert "lifecycle columns missing" in response.json()["detail"]


@pytest.mark.asyncio
async def test_list_and_get_product_draft():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        tenant_header = {"x-shop-domain": "test-shop.myshopify.com"}
        payload = {
            "products_json": json.dumps([{"title": "Draft A"}]),
            "run_id": "run-a",
            "import_mode": "create",
            "shop_domain": tenant_header["x-shop-domain"],
            "input_file_id": "input-file-a",
            "input_filename": "draft-input.xlsx",
        }
        created = await ac.post("/agents/product-drafts", data=payload)
        assert created.status_code == 200
        draft_id = created.json()["draft_id"]

        listed = await ac.get("/agents/product-drafts", headers=tenant_header)
        assert listed.status_code == 200
        drafts = listed.json()["drafts"]
        assert any(d.get("draft_id") == draft_id for d in drafts)

        detail = await ac.get(f"/agents/product-drafts/{draft_id}")
        assert detail.status_code == 200
        assert detail.json()["draft"]["draft_id"] == draft_id
        assert detail.json()["draft"]["input_file_id"] == "input-file-a"
        assert detail.json()["draft"]["input_filename"] == "draft-input.xlsx"

        resume = await ac.post(f"/agents/product-drafts/{draft_id}/resume-file")
        assert resume.status_code == 200
        resume_body = resume.json()
        assert "file_id" in resume_body
        assert resume_body["filename"].endswith(".xlsx")
        persisted_detail = await ac.get(f"/agents/product-drafts/{draft_id}")
        assert persisted_detail.status_code == 200
        persisted_draft = persisted_detail.json()["draft"]
        assert persisted_draft["output_file_id"] == resume_body["file_id"]
        assert persisted_draft["output_filename"] == resume_body["filename"]
        resume_again = await ac.post(f"/agents/product-drafts/{draft_id}/resume-file")
        assert resume_again.status_code == 200
        assert resume_again.json()["file_id"] == resume_body["file_id"]


@pytest.mark.asyncio
async def test_product_drafts_list_requires_shop_domain_header():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.get("/agents/product-drafts")
    assert response.status_code == 400
    assert "shop_domain" in response.text


@pytest.mark.asyncio
async def test_product_drafts_list_accepts_shop_query_param():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        created = await ac.post(
            "/agents/product-drafts",
            data={
                "products_json": json.dumps([{"title": "Query Shop Draft"}]),
                "run_id": "run-query-shop",
                "import_mode": "auto",
                "shop_domain": "query-shop.myshopify.com",
            },
        )
        assert created.status_code == 200
        draft_id = created.json()["draft_id"]

        listed = await ac.get("/agents/product-drafts?shop=query-shop.myshopify.com")
        assert listed.status_code == 200
        draft_ids = [item.get("draft_id") for item in listed.json()["drafts"]]
        assert draft_id in draft_ids


@pytest.mark.asyncio
async def test_submitted_documents_list_requires_shop_domain_header():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        response = await ac.get("/agents/submitted-documents")
    assert response.status_code == 400
    assert "shop_domain" in response.text


@pytest.mark.asyncio
async def test_bulk_delete_product_drafts():
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        create_payload = {
            "products_json": json.dumps([{"title": "Bulk Draft"}]),
            "run_id": "run-bulk-drafts",
            "import_mode": "create",
        }
        created = await ac.post("/agents/product-drafts", data=create_payload)
        assert created.status_code == 200
        draft_id = created.json()["draft_id"]

        missing_id = "missing-draft-id"
        bulk_response = await ac.post(
            "/agents/product-drafts/bulk-delete",
            json={"ids": [draft_id, missing_id]},
        )
        assert bulk_response.status_code == 200
        body = bulk_response.json()
        assert body["deleted_ids"] == [draft_id]
        assert body["failed_ids"] == [missing_id]
        assert (await ac.get(f"/agents/product-drafts/{draft_id}")).status_code == 404


@pytest.mark.asyncio
async def test_successful_submit_creates_submitted_and_hides_draft(monkeypatch):
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {
                    "id": "gid://shopify/Product/1",
                    "title": "Submitted Draft Product",
                },
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        tenant_header = {"x-shop-domain": "test-shop.myshopify.com"}
        draft_payload = {
            "products_json": json.dumps([{"title": "Submitted Draft Product"}]),
            "run_id": "run-submit",
            "import_mode": "create",
            "shop_domain": tenant_header["x-shop-domain"],
            "draft_name": "submitted-draft.xlsx",
            "input_file_id": "input-file-preview",
            "input_filename": "submitted-source.xlsx",
        }
        created_draft = await ac.post("/agents/product-drafts", data=draft_payload)
        assert created_draft.status_code == 200
        draft_id = created_draft.json()["draft_id"]

        submit_payload = {
            "products_json": json.dumps([{"title": "Submitted Draft Product"}]),
            "import_mode": "create",
            "run_id": "run-submit-1",
            "draft_id": draft_id,
            "document_name": "submitted-draft.xlsx",
            "shop_domain": tenant_header["x-shop-domain"],
        }
        submitted = await ac.post("/agents/submit-products", data=submit_payload)
        assert submitted.status_code == 200
        submitted_body = submitted.json()
        assert submitted_body["success_count"] == 1
        assert submitted_body["submitted_id"]

        drafts_after_submit = await ac.get(
            "/agents/product-drafts",
            headers=tenant_header,
        )
        assert drafts_after_submit.status_code == 200
        draft_ids = [d.get("draft_id") for d in drafts_after_submit.json()["drafts"]]
        assert draft_id not in draft_ids

        submitted_list = await ac.get(
            "/agents/submitted-documents",
            headers=tenant_header,
        )
        assert submitted_list.status_code == 200
        items = submitted_list.json()["submitted_documents"]
        assert any(
            item.get("submitted_id") == submitted_body["submitted_id"] for item in items
        )
        matching_item = next(
            item
            for item in items
            if item.get("submitted_id") == submitted_body["submitted_id"]
        )
        assert matching_item.get("preview_file_id") == "input-file-preview"

        submitted_detail = await ac.get(
            f"/agents/submitted-documents/{submitted_body['submitted_id']}"
        )
        assert submitted_detail.status_code == 200

        submitted_resume = await ac.post(
            f"/agents/submitted-documents/{submitted_body['submitted_id']}/resume-file"
        )
        assert submitted_resume.status_code == 200
        assert submitted_resume.json()["filename"].endswith(".xlsx")


@pytest.mark.asyncio
async def test_bulk_delete_submitted_documents(monkeypatch):
    _patch_bulk_submit(monkeypatch, [{
        "data": {
            "productSet": {
                "product": {
                    "id": "gid://shopify/Product/3",
                    "title": "Bulk Submitted",
                },
                "userErrors": [],
            }
        }
    }])

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://testserver") as ac:
        submit_payload = {
            "products_json": json.dumps([{"title": "Bulk Submitted"}]),
            "import_mode": "create",
            "run_id": "run-bulk-submitted",
            "document_name": "bulk-submitted.xlsx",
        }
        created_submitted = await ac.post(
            "/agents/submit-products", data=submit_payload
        )
        assert created_submitted.status_code == 200
        submitted_id = created_submitted.json()["submitted_id"]

        missing_id = "missing-submitted-id"
        bulk_response = await ac.post(
            "/agents/submitted-documents/bulk-delete",
            json={"ids": [submitted_id, missing_id]},
        )
        assert bulk_response.status_code == 200
        body = bulk_response.json()
        assert body["deleted_ids"] == [submitted_id]
        assert body["failed_ids"] == [missing_id]
        assert (
            await ac.get(f"/agents/submitted-documents/{submitted_id}")
        ).status_code == 404


def _make_png_bytes(mode: str, size: tuple[int, int], color) -> bytes:
    image = Image.new(mode, size, color)
    output = io.BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def test_is_blank_png_detects_transparent_and_white():
    transparent_png = _make_png_bytes("RGBA", (8, 8), (0, 0, 0, 0))
    white_png = _make_png_bytes("RGBA", (8, 8), (255, 255, 255, 255))
    black_png = _make_png_bytes("RGBA", (8, 8), (0, 0, 0, 255))

    assert _is_blank_png(transparent_png) is True
    assert _is_blank_png(white_png) is True
    assert _is_blank_png(black_png) is False


@pytest.mark.asyncio
async def test_generate_thumbnail_bytes_skips_blank_pages():
    blank_transparent = _make_png_bytes("RGBA", (8, 8), (0, 0, 0, 0))
    blank_white = _make_png_bytes("RGBA", (8, 8), (255, 255, 255, 255))
    non_blank = _make_png_bytes("RGBA", (8, 8), (0, 0, 0, 255))

    class _FakeCollabora:
        async def convert_document_to_png_collabora(self, *args, **kwargs):
            _ = (args, kwargs)
            return [blank_transparent, blank_white, non_blank]

    selected = await _generate_thumbnail_bytes(
        file_bytes=b"dummy",
        filename="doc.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        collabora_url="http://localhost:8080",
        collabora=_FakeCollabora(),
    )
    assert selected == non_blank


@pytest.mark.asyncio
async def test_run_and_get_product_intelligence_audit(monkeypatch):
    async def fake_get_product(self, gid):
        return {
            "data": {
                "node": {
                    "id": gid,
                    "title": "Short",
                    "descriptionHtml": "",
                    "vendor": "Brand",
                    "handle": "short",
                    "productType": "General",
                    "status": "ACTIVE",
                    "tags": ["t1"],
                    "seo": {"title": "Short", "description": "Desc"},
                }
            }
        }

    async def fake_update_product_from_input(self, product):
        return {
            "data": {
                "productUpdate": {
                    "product": {
                        "id": product.get("id"),
                        "title": product.get("title"),
                    },
                    "userErrors": [],
                }
            }
        }

    async def fake_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = (supabase, shop_domain, normalization_settings, trace_event)
        target_title = str(products[0].get("title") or "Short") if products else "Short"
        return [
            {
                "suggestion_id": f"suggestion-{uuid.uuid4()}",
                "finding_id": f"finding-{uuid.uuid4()}",
                "product_index": 0,
                "product_title": target_title,
                "category": "completeness",
                "severity": "medium",
                "message": "Add vendor detail",
                "patch_payload": {"vendor": "Improved Vendor"},
                "status": "pending",
            }
        ]

    import shopify as shopify_module
    import application.use_cases.intelligence_generate_suggestions as suggestions_uc

    monkeypatch.setattr(
        shopify_module.ShopifyClient,
        "update_product_from_input",
        fake_update_product_from_input,
    )
    monkeypatch.setattr(shopify_module.ShopifyClient, "get_product", fake_get_product)
    monkeypatch.setattr(
        suggestions_uc, "execute", fake_generate_suggestions_execute
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(
        transport=transport, base_url="http://testserver", headers=TEST_SHOP_HEADERS
    ) as ac:
        audit_run = await ac.post(
            "/agents/intelligence/audit",
            json={
                "products": [
                    {
                        "id": "gid://shopify/Product/9",
                        "title": "Short",
                        "handle": "short",
                    }
                ]
            },
        )
        assert audit_run.status_code == 200
        audit = audit_run.json()
        assert audit["audit_id"]
        assert isinstance(audit.get("overall_score"), int)
        assert isinstance(audit.get("findings"), list)

        audits = await ac.get("/agents/intelligence/audits")
        assert audits.status_code == 200
        listed = audits.json()["audits"]
        assert any(item.get("audit_id") == audit["audit_id"] for item in listed)

        detail = await ac.get(f"/agents/intelligence/audits/{audit['audit_id']}")
        assert detail.status_code == 200
        assert detail.json()["audit"]["audit_id"] == audit["audit_id"]

        suggestions = await ac.get(
            f"/agents/intelligence/audits/{audit['audit_id']}/suggestions"
        )
        assert suggestions.status_code == 200
        suggestion_rows = suggestions.json()["suggestions"]
        assert isinstance(suggestion_rows, list)
        assert suggestion_rows

        suggestion_id = suggestion_rows[0]["suggestion_id"]
        apply_single = await ac.post(
            f"/agents/intelligence/suggestions/{suggestion_id}/apply"
        )
        assert apply_single.status_code == 200
        assert apply_single.json()["status"] == "applied"
        assert apply_single.json()["shopify_updated"] is True
        assert apply_single.json()["target_product_id"] == "gid://shopify/Product/9"

        revert_single = await ac.post(
            f"/agents/intelligence/suggestions/{suggestion_id}/revert"
        )
        assert revert_single.status_code == 200
        assert revert_single.json()["status"] == "pending"
        assert revert_single.json()["shopify_updated"] is True
        assert revert_single.json()["target_product_id"] == "gid://shopify/Product/9"

        remaining_ids = [
            row["suggestion_id"]
            for row in suggestion_rows[1:]
            if row.get("status") != "applied"
        ]
        if remaining_ids:
            apply_bulk = await ac.post(
                "/agents/intelligence/suggestions/apply-bulk",
                json={"suggestion_ids": remaining_ids},
            )
            assert apply_bulk.status_code == 200
            assert apply_bulk.json()["failed_count"] == 0


@pytest.mark.asyncio
async def test_intelligence_audit_with_existing_products_mode(monkeypatch):
    async def fake_list_products_for_audit(self, query=None, limit=50):
        _ = (self, query, limit)
        return [
            {
                "id": "gid://shopify/Product/101",
                "title": "Catalog Product",
                "handle": "catalog-product",
                "vendor": "Brand",
                "product_type": "General",
                "status": "active",
                "tags": ["tag-1"],
                "body_html": "Catalog description",
                "seo_title": "Catalog Product",
                "seo_description": "Catalog description",
                "variants": [{"sku": "SKU-101"}],
            }
        ]

    async def fake_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = (supabase, shop_domain, normalization_settings, trace_event)
        target_title = (
            str(products[0].get("title") or "Catalog Product")
            if products
            else "Catalog Product"
        )
        return [
            {
                "suggestion_id": f"suggestion-{uuid.uuid4()}",
                "finding_id": f"finding-{uuid.uuid4()}",
                "product_index": 0,
                "product_title": target_title,
                "category": "seo_readiness",
                "severity": "low",
                "message": "Improve vendor detail",
                "patch_payload": {"vendor": "Catalog Vendor"},
                "status": "pending",
            }
        ]

    import infrastructure.adapters.shopify_adapter as shopify_adapter
    import application.use_cases.intelligence_generate_suggestions as suggestions_uc

    monkeypatch.setattr(
        shopify_adapter.ShopifyAdapter,
        "list_products_for_audit",
        fake_list_products_for_audit,
    )
    monkeypatch.setattr(
        suggestions_uc, "execute", fake_generate_suggestions_execute
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(
        transport=transport, base_url="http://testserver", headers=TEST_SHOP_HEADERS
    ) as ac:
        products_response = await ac.get(
            "/agents/intelligence/shopify-products?query=catalog&limit=10"
        )
        assert products_response.status_code == 200
        products = products_response.json()["products"]
        assert len(products) == 1
        assert products[0]["title"] == "Catalog Product"

        selected_audit = await ac.post(
            "/agents/intelligence/audit",
            json={"products": products},
        )
        assert selected_audit.status_code == 200
        assert selected_audit.json()["audit_id"]

        all_audit = await ac.post(
            "/agents/intelligence/audit",
            json={"all_products": True, "query": "catalog", "limit": 10},
        )
        assert all_audit.status_code == 200
        assert all_audit.json()["audit_id"]


@pytest.mark.asyncio
async def test_apply_suggestion_allows_missing_previous_values(monkeypatch):
    async def fake_get_product(self, gid):
        return {
            "data": {
                "node": {
                    "id": gid,
                    "title": "Catalog Product",
                    "descriptionHtml": "",
                    "vendor": None,
                    "handle": "catalog-product",
                    "productType": "General",
                    "status": "ACTIVE",
                    "tags": ["tag-1"],
                    "seo": {
                        "title": "Catalog Product",
                        "description": "Catalog description",
                    },
                }
            }
        }

    updates: list[dict[str, object]] = []

    async def fake_update_product_from_input(self, product):
        updates.append(dict(product))
        return {
            "data": {
                "productUpdate": {
                    "product": {"id": product.get("id")},
                    "userErrors": [],
                }
            }
        }

    async def fake_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = (supabase, shop_domain, normalization_settings, trace_event)
        target_title = (
            str(products[0].get("title") or "Catalog Product")
            if products
            else "Catalog Product"
        )
        return [
            {
                "suggestion_id": f"suggestion-{uuid.uuid4()}",
                "finding_id": f"finding-{uuid.uuid4()}",
                "product_index": 0,
                "product_title": target_title,
                "category": "completeness",
                "severity": "medium",
                "message": "Add missing vendor",
                "patch_payload": {"vendor": "Improved Vendor"},
                "status": "pending",
            }
        ]

    import shopify as shopify_module
    import application.use_cases.intelligence_generate_suggestions as suggestions_uc

    monkeypatch.setattr(shopify_module.ShopifyClient, "get_product", fake_get_product)
    monkeypatch.setattr(
        shopify_module.ShopifyClient,
        "update_product_from_input",
        fake_update_product_from_input,
    )
    monkeypatch.setattr(
        suggestions_uc, "execute", fake_generate_suggestions_execute
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(
        transport=transport, base_url="http://testserver", headers=TEST_SHOP_HEADERS
    ) as ac:
        audit_run = await ac.post(
            "/agents/intelligence/audit",
            json={
                "products": [
                    {
                        "id": "gid://shopify/Product/201",
                        "title": "Catalog Product",
                        "handle": "catalog-product",
                    }
                ]
            },
        )
        assert audit_run.status_code == 200
        audit_id = audit_run.json()["audit_id"]

        suggestions = await ac.get(
            f"/agents/intelligence/audits/{audit_id}/suggestions"
        )
        assert suggestions.status_code == 200
        rows = suggestions.json()["suggestions"]
        suggestion = next(
            (
                row
                for row in rows
                if isinstance(row.get("patch_payload"), dict)
                and "vendor" in row["patch_payload"]
            ),
            None,
        )
        assert suggestion is not None

        apply_single = await ac.post(
            f"/agents/intelligence/suggestions/{suggestion['suggestion_id']}/apply"
        )
        assert apply_single.status_code == 200
        assert apply_single.json()["status"] == "applied"
        applied_suggestion = apply_single.json()["suggestion"]
        assert applied_suggestion["previous_payload"]["vendor"] is None
        assert applied_suggestion["previous_payload"]["__is_reversible"] is True
        assert (
            applied_suggestion["previous_payload"]["__revert_modes"]["vendor"]
            == "clear"
        )

        revert_single = await ac.post(
            f"/agents/intelligence/suggestions/{suggestion['suggestion_id']}/revert"
        )
        assert revert_single.status_code == 200
        assert revert_single.json()["status"] == "pending"
        assert len(updates) == 2
        assert updates[1]["vendor"] == ""


@pytest.mark.asyncio
async def test_apply_partial_field_keeps_remaining_fields_pending(monkeypatch):
    async def fake_get_product(self, gid):
        return {
            "data": {
                "node": {
                    "id": gid,
                    "title": "Catalog Product",
                    "descriptionHtml": "Catalog description",
                    "vendor": "Original Vendor",
                    "handle": "catalog-product",
                    "productType": "General",
                    "status": "ACTIVE",
                    "tags": ["tag-1"],
                    "seo": {
                        "title": "Original SEO",
                        "description": "Original description",
                    },
                }
            }
        }

    updates: list[dict[str, object]] = []

    async def fake_update_product_from_input(self, product):
        updates.append(dict(product))
        return {
            "data": {
                "productUpdate": {
                    "product": {"id": product.get("id")},
                    "userErrors": [],
                }
            }
        }

    import shopify as shopify_module

    monkeypatch.setattr(shopify_module.ShopifyClient, "get_product", fake_get_product)
    monkeypatch.setattr(
        shopify_module.ShopifyClient,
        "update_product_from_input",
        fake_update_product_from_input,
    )

    ctx = get_app_context()
    audit_id = f"audit-partial-{uuid.uuid4()}"
    suggestion_id = f"suggestion-partial-{uuid.uuid4()}"
    ctx.services.supabase.intelligence.save_product_intelligence_audit(
        audit_id=audit_id,
        run_id=None,
        submitted_id=None,
        scope="adhoc_products",
        status="success",
        overall_score=60,
        findings_count=1,
        component_scores={},
        totals={
            "audited_products": [
                {
                    "id": "gid://shopify/Product/401",
                    "title": "Catalog Product",
                    "handle": "catalog-product",
                }
            ]
        },
        shop_domain=TEST_SHOP_DOMAIN,
    )
    ctx.services.supabase.intelligence.save_product_intelligence_suggestions(
        audit_id=audit_id,
        suggestions=[
            {
                "suggestion_id": suggestion_id,
                "finding_id": f"finding-{uuid.uuid4()}",
                "product_index": 0,
                "product_title": "Catalog Product",
                "category": "seo_readiness",
                "severity": "medium",
                "message": "Apply vendor and SEO title improvements",
                "patch_payload": {
                    "vendor": "Updated Vendor",
                    "seo_title": "Updated SEO Title",
                },
                "status": "pending",
            }
        ],
        shop_domain=TEST_SHOP_DOMAIN,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(
        transport=transport, base_url="http://testserver", headers=TEST_SHOP_HEADERS
    ) as ac:
        apply_single = await ac.post(
            f"/agents/intelligence/suggestions/{suggestion_id}/apply",
            json={"patch_payload": {"vendor": "Updated Vendor"}},
        )
        assert apply_single.status_code == 200
        assert apply_single.json()["status"] == "applied"
        assert apply_single.json()["suggestion"]["patch_payload"] == {
            "vendor": "Updated Vendor"
        }
        assert len(updates) == 1
        assert updates[0]["vendor"] == "Updated Vendor"
        assert "seo_title" not in updates[0]

        suggestions = await ac.get(
            f"/agents/intelligence/audits/{audit_id}/suggestions"
        )
        assert suggestions.status_code == 200
        rows = suggestions.json()["suggestions"]
        applied = [row for row in rows if row.get("status") == "applied"]
        pending = [row for row in rows if row.get("status") != "applied"]
        assert len(applied) == 1
        assert len(pending) == 1
        assert pending[0]["patch_payload"] == {"seo_title": "Updated SEO Title"}


@pytest.mark.asyncio
async def test_revert_blocked_for_non_reversible_missing_original_field(monkeypatch):
    async def fake_get_product(self, gid):
        return {
            "data": {
                "node": {
                    "id": gid,
                    "title": "Catalog Product",
                    "descriptionHtml": "",
                    "vendor": "Brand",
                    "handle": "catalog-product",
                    "productType": "General",
                    "status": "ACTIVE",
                    "tags": ["tag-1"],
                    "seo": {
                        "title": "Catalog Product",
                        "description": "Catalog description",
                    },
                }
            }
        }

    async def fake_update_product_from_input(self, product):
        _ = product
        return {
            "data": {
                "productUpdate": {
                    "product": {"id": "gid://shopify/Product/301"},
                    "userErrors": [],
                }
            }
        }

    import shopify as shopify_module

    monkeypatch.setattr(shopify_module.ShopifyClient, "get_product", fake_get_product)
    monkeypatch.setattr(
        shopify_module.ShopifyClient,
        "update_product_from_input",
        fake_update_product_from_input,
    )

    ctx = get_app_context()
    audit_id = f"audit-nonreversible-{uuid.uuid4()}"
    suggestion_id = f"suggestion-nonreversible-{uuid.uuid4()}"
    ctx.services.supabase.intelligence.save_product_intelligence_audit(
        audit_id=audit_id,
        run_id=None,
        submitted_id=None,
        scope="adhoc_products",
        status="success",
        overall_score=60,
        findings_count=1,
        component_scores={},
        totals={
            "audited_products": [
                {
                    "id": "gid://shopify/Product/301",
                    "title": "Catalog Product",
                    "handle": "catalog-product",
                }
            ]
        },
        shop_domain=TEST_SHOP_DOMAIN,
    )
    ctx.services.supabase.intelligence.save_product_intelligence_suggestions(
        audit_id=audit_id,
        suggestions=[
            {
                "suggestion_id": suggestion_id,
                "finding_id": f"finding-{uuid.uuid4()}",
                "product_index": 0,
                "product_title": "Catalog Product",
                "category": "consistency",
                "severity": "low",
                "message": "Apply unsupported field fix",
                "patch_payload": {"nonexistent_field": "value"},
                "status": "pending",
            }
        ],
        shop_domain=TEST_SHOP_DOMAIN,
    )

    transport = ASGITransport(app=app)
    async with AsyncClient(
        transport=transport, base_url="http://testserver", headers=TEST_SHOP_HEADERS
    ) as ac:
        apply_single = await ac.post(
            f"/agents/intelligence/suggestions/{suggestion_id}/apply"
        )
        assert apply_single.status_code == 200
        applied_suggestion = apply_single.json()["suggestion"]
        assert applied_suggestion["previous_payload"]["__is_reversible"] is False

        revert_single = await ac.post(
            f"/agents/intelligence/suggestions/{suggestion_id}/revert"
        )
        assert revert_single.status_code == 400
        assert "Auto-revert is unavailable" in revert_single.json()["detail"]
