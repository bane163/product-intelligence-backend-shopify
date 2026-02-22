import io

import openpyxl

from ai.excel_writer import create_excel_bytes
from ai.models import ProductInput, ProductMetafield, ProductsList


def test_create_excel_bytes_includes_generated_fields_and_ai_sheet():
    products = ProductsList(
        products=[
            ProductInput(
                title="Enhanced Product",
                body_html="<p>Enhanced description</p>",
                vendor="Enhanced Vendor",
                product_category="Apparel",
                product_type="Shirt",
                tags=["summer", "cotton"],
                seo_title="Enhanced SEO Title",
                seo_description="Enhanced SEO Description",
                status="draft",
                published=True,
                metafields=[
                    ProductMetafield(
                        namespace="extractor",
                        key="source_confidence",
                        value="high",
                        type="single_line_text_field",
                    )
                ],
            )
        ]
    )

    workbook_bytes = create_excel_bytes(products)
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))

    assert "Products" in workbook.sheetnames
    assert "AI Enhancements" in workbook.sheetnames

    products_sheet = workbook["Products"]
    headers = [cell.value for cell in products_sheet[1]]
    first_row = [cell.value for cell in products_sheet[2]]
    data = {str(header): first_row[index] for index, header in enumerate(headers)}

    assert data["Product Category"] == "Apparel"
    assert data["Type"] == "Shirt"
    assert data["Tags"] == "summer, cotton"
    assert data["SEO Title"] == "Enhanced SEO Title"
    assert data["SEO Description"] == "Enhanced SEO Description"
    assert data["Status"] == "draft"

    ai_sheet = workbook["AI Enhancements"]
    ai_headers = [cell.value for cell in ai_sheet[1]]
    ai_row = [cell.value for cell in ai_sheet[2]]
    ai_data = {str(header): ai_row[index] for index, header in enumerate(ai_headers)}
    assert ai_data["Attribute"] == "metafield:extractor.source_confidence"
    assert ai_data["Value"] == "high"
    assert ai_data["Type"] == "single_line_text_field"
