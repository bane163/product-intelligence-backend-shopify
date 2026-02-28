from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

import application.use_cases.processing.process_document as process_document_uc


class _FakeRuns:
    def __init__(self) -> None:
        self.created_updates: list[tuple[str, dict]] = []
        self.finalized: list[tuple[str, dict]] = []
        self.events: list[tuple[str, dict, int]] = []
        self.messages: list[tuple[str, str, str, int, dict | None]] = []

    def create_or_update_run(self, run_id: str, fields: dict) -> None:
        self.created_updates.append((run_id, dict(fields)))

    def finalize_run(
        self,
        run_id: str,
        *,
        status: str,
        duration_ms: int | None = None,
        error: str | None = None,
        extra_fields: dict | None = None,
    ) -> None:
        self.finalized.append(
            (
                run_id,
                {
                    "status": status,
                    "duration_ms": duration_ms,
                    "error": error,
                    "extra_fields": dict(extra_fields or {}),
                },
            )
        )

    def append_run_event(self, run_id: str, event: dict, seq: int) -> None:
        self.events.append((run_id, dict(event), seq))

    def append_run_message(
        self,
        run_id: str,
        *,
        role: str,
        message: str,
        seq: int,
        meta: dict | None = None,
    ) -> None:
        self.messages.append((run_id, role, message, seq, meta))


class _FakeLLMConfigs:
    def __init__(self, config: dict | None) -> None:
        self._config = config

    def get_active_llm_model_config(self, shop_domain: str) -> dict | None:
        _ = shop_domain
        return self._config


class _FakeSupabase:
    def __init__(self, model_config: dict | None = None) -> None:
        self.runs = _FakeRuns()
        self.llm_configs = _FakeLLMConfigs(model_config)
        self.file = _FakeFileStorage()


class _FakeFileStorage:
    def __init__(self) -> None:
        self.saved_files: list[dict] = []

    def save_file(
        self,
        file_id: str,
        name: str,
        content: bytes,
        *,
        content_type: str,
        file_origin: str,
    ) -> None:
        self.saved_files.append(
            {
                "file_id": file_id,
                "name": name,
                "content": content,
                "content_type": content_type,
                "file_origin": file_origin,
            }
        )


class _FakeTracing:
    def __init__(self) -> None:
        self.completed_runs: list[str] = []
        self.events: list[dict] = []

    def emit_run_event(
        self,
        run_id: str,
        *,
        phase: str,
        message: str,
        level: str = "info",
        payload_preview=None,
        error: str | None = None,
        metadata: dict | None = None,
    ) -> dict:
        event = {
            "run_id": run_id,
            "phase": phase,
            "message": message,
            "level": level,
            "payload_preview": payload_preview,
            "error": error,
            "metadata": metadata,
        }
        self.events.append(event)
        return event

    def complete_run(self, run_id: str) -> None:
        self.completed_runs.append(run_id)


class _FakeLLM:
    def __init__(self, payload: dict) -> None:
        self.payload = payload

    async def run_excel_agent_workflow(self, *args, **kwargs):
        _ = args, kwargs
        return dict(self.payload)


@pytest.mark.asyncio
async def test_process_document_enriches_products_during_import(monkeypatch):
    async def fake_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = supabase, shop_domain, normalization_settings, trace_event
        assert products == [{"title": "Demo Product"}]
        return [
            {
                "product_index": 0,
                "patch_payload": {"vendor": "Enriched Vendor"},
            }
        ]

    monkeypatch.setattr(
        process_document_uc,
        "generate_suggestions_execute",
        fake_generate_suggestions_execute,
        raising=False,
    )

    supabase = _FakeSupabase(
        model_config={
            "base_url": "http://localhost:11434/v1/",
            "model_id": "deepseek-r1:8b",
            "api_key": "secret",
            "provider": "ollama",
        }
    )
    llm = _FakeLLM(payload={"products": [{"title": "Demo Product"}]})
    tracing = _FakeTracing()

    result = await process_document_uc.execute(
        supabase=supabase,
        llm=llm,
        tracing=tracing,
        ctx=SimpleNamespace(),
        file_bytes=b"sheet-bytes",
        input_name="catalog.xlsx",
        input_content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        run_id="run-import-enrich",
        shop_domain="store.myshopify.com",
    )

    assert result["result"]["products"][0]["vendor"] == "Enriched Vendor"
    assert result["result"]["enrichment_applied"] is True
    assert result["result"]["enrichment_suggestions_count"] == 1


@pytest.mark.asyncio
async def test_process_document_refreshes_workbook_output_after_enrichment(monkeypatch):
    async def fake_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = supabase, products, shop_domain, normalization_settings, trace_event
        return [
            {
                "product_index": 0,
                "patch_payload": {"vendor": "Enriched Vendor"},
            }
        ]

    monkeypatch.setattr(
        process_document_uc,
        "generate_suggestions_execute",
        fake_generate_suggestions_execute,
        raising=False,
    )

    supabase = _FakeSupabase(
        model_config={
            "base_url": "http://localhost:11434/v1/",
            "model_id": "deepseek-r1:8b",
            "api_key": "secret",
            "provider": "ollama",
        }
    )
    llm = _FakeLLM(
        payload={
            "file_id": "pre-enrichment-file",
            "filename": "catalog-products.xlsx",
            "products": [{"title": "Demo Product"}],
        }
    )
    tracing = _FakeTracing()

    result = await process_document_uc.execute(
        supabase=supabase,
        llm=llm,
        tracing=tracing,
        ctx=SimpleNamespace(),
        file_bytes=b"sheet-bytes",
        input_name="catalog.xlsx",
        input_content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        run_id="run-import-refresh-output",
        shop_domain="store.myshopify.com",
        write_to_file=True,
    )

    payload = result["result"]
    assert payload["products"][0]["vendor"] == "Enriched Vendor"
    assert payload["file_id"] != "pre-enrichment-file"
    assert len(supabase.file.saved_files) == 1
    saved = supabase.file.saved_files[0]
    assert saved["file_id"] == payload["file_id"]
    assert saved["name"] == payload["filename"]
    finalized_extra = supabase.runs.finalized[-1][1]["extra_fields"]
    assert finalized_extra["output_file_id"] == payload["file_id"]
    assert finalized_extra["output_filename"] == payload["filename"]
    workbook = load_workbook(BytesIO(saved["content"]))
    worksheet = workbook["Products"]
    headers = {str(cell.value): index + 1 for index, cell in enumerate(worksheet[1])}
    assert worksheet.cell(row=2, column=headers["Vendor"]).value == "Enriched Vendor"


@pytest.mark.asyncio
async def test_process_document_enrichment_failure_is_non_fatal(monkeypatch):
    async def fail_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = supabase, products, shop_domain, normalization_settings, trace_event
        raise RuntimeError("Synthetic enrichment failure")

    monkeypatch.setattr(
        process_document_uc,
        "generate_suggestions_execute",
        fail_generate_suggestions_execute,
        raising=False,
    )

    supabase = _FakeSupabase(
        model_config={
            "base_url": "http://localhost:11434/v1/",
            "model_id": "deepseek-r1:8b",
            "api_key": "secret",
            "provider": "ollama",
        }
    )
    llm = _FakeLLM(payload={"products": [{"title": "Demo Product"}]})
    tracing = _FakeTracing()

    result = await process_document_uc.execute(
        supabase=supabase,
        llm=llm,
        tracing=tracing,
        ctx=SimpleNamespace(),
        file_bytes=b"sheet-bytes",
        input_name="catalog.xlsx",
        input_content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        run_id="run-import-enrich-fallback",
        shop_domain="store.myshopify.com",
    )

    assert result["result"]["products"][0]["title"] == "Demo Product"
    assert "vendor" not in result["result"]["products"][0]
    assert result["result"]["enrichment_applied"] is False
    assert "Synthetic enrichment failure" in result["result"]["enrichment_warning"]


@pytest.mark.asyncio
async def test_process_document_workbook_ai_enhancements_shows_all_enriched_fields(monkeypatch):
    """The AI Enhancements sheet must list ALL fields changed by enrichment,
    not just metafields.  Non-metafield patches (vendor, seo_title, tags, etc.)
    should appear alongside metafield patches."""

    async def fake_generate_suggestions_execute(
        *,
        supabase,
        products,
        shop_domain,
        normalization_settings=None,
        trace_event=None,
    ):
        _ = supabase, products, shop_domain, normalization_settings, trace_event
        return [
            {
                "product_index": 0,
                "patch_payload": {
                    "vendor": "Enriched Vendor",
                    "seo_title": "Better SEO Title",
                    "tags": "tag-a, tag-b",
                    "metafields": [
                        {
                            "namespace": "extractor",
                            "key": "color_hint",
                            "value": "blue",
                            "type": "single_line_text_field",
                        }
                    ],
                },
            }
        ]

    monkeypatch.setattr(
        process_document_uc,
        "generate_suggestions_execute",
        fake_generate_suggestions_execute,
        raising=False,
    )

    supabase = _FakeSupabase(
        model_config={
            "base_url": "http://localhost:11434/v1/",
            "model_id": "deepseek-r1:8b",
            "api_key": "secret",
            "provider": "ollama",
        }
    )
    llm = _FakeLLM(
        payload={
            "file_id": "pre-enrichment-file",
            "filename": "catalog-products.xlsx",
            "products": [{"title": "Demo Product"}],
        }
    )
    tracing = _FakeTracing()

    result = await process_document_uc.execute(
        supabase=supabase,
        llm=llm,
        tracing=tracing,
        ctx=SimpleNamespace(),
        file_bytes=b"sheet-bytes",
        input_name="catalog.xlsx",
        input_content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        run_id="run-ai-enhancements-sheet",
        shop_domain="store.myshopify.com",
        write_to_file=True,
    )

    saved = supabase.file.saved_files[0]
    workbook = load_workbook(BytesIO(saved["content"]))
    ai_sheet = workbook["AI Enhancements"]

    # Collect all data rows (skip header)
    rows = []
    for row in ai_sheet.iter_rows(min_row=2, values_only=True):
        attr = row[2]  # Attribute column
        if attr and attr != "No AI-only attributes captured":
            rows.append({"attribute": attr, "value": row[3], "type": row[4]})

    attributes = [r["attribute"] for r in rows]
    # Non-metafield enrichments must appear
    assert "vendor" in attributes, f"vendor missing from AI Enhancements: {attributes}"
    assert "seo_title" in attributes, f"seo_title missing from AI Enhancements: {attributes}"
    assert "tags" in attributes, f"tags missing from AI Enhancements: {attributes}"
    # Metafield enrichments must still appear
    assert "metafield:extractor.color_hint" in attributes, (
        f"metafield missing from AI Enhancements: {attributes}"
    )
