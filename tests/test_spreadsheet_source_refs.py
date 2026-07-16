import io

from openpyxl import Workbook

from application.services.spreadsheet_source_refs import enrich_spreadsheet_source_refs


def workbook_bytes(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def multisheet_workbook_bytes(sheets):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_adds_exact_refs_for_first_variant_and_normalizes_numeric_price():
    products = [{"title": "Lamp", "vendor": "Acme", "variants": [{"sku": "L-1", "price": "12.00"}]}]
    data = workbook_bytes([
        ["Product Name", "Vendor", "Variant SKU", "Retail Price"],
        ["Other", "Else", "X", 4],
        ["Lamp", "Acme", "L-1", 12],
    ])
    enrich_spreadsheet_source_refs(products, data, "products.xlsx")
    assert {(ref["field"], ref["cell"]) for ref in products[0]["source_refs"]} == {
        ("title", "A3"), ("vendor", "B3"), ("sku", "C3"), ("price", "D3")
    }


def test_replaces_invalid_existing_sku_ref_on_anchored_sheet():
    existing = {"field": "variants[0].sku", "sheet": "Products", "cell": "Z9"}
    products = [{"title": "Lamp", "variants": [{"sku": "L-1", "price": "12"}], "source_refs": [
        {"field": "title", "sheet": "Products", "cell": "A2"}, existing,
    ]}]
    enrich_spreadsheet_source_refs(products, workbook_bytes([["Title", "SKU", "Price"], ["Lamp", "L-1", 12]]), "p.xlsx")
    assert existing not in products[0]["source_refs"]
    assert any(ref.get("field") == "sku" and ref.get("sheet") == "Products" and ref.get("cell") == "B2" for ref in products[0]["source_refs"])


def test_refuses_ambiguous_duplicate_product_rows():
    products = [{"title": "Lamp", "variants": [{"price": "12"}]}]
    enrich_spreadsheet_source_refs(products, workbook_bytes([["Title", "Price"], ["Lamp", 12], ["Lamp", 12]]), "p.xlsx")
    assert products[0]["source_refs"] == []


def test_unique_sku_selects_correct_row_when_title_is_duplicated():
    products = [{"title": "Lamp", "vendor": "Acme", "variants": [{"sku": "B", "price": 20}]}]
    enrich_spreadsheet_source_refs(products, workbook_bytes([
        ["Title", "Brand", "Product Code", "Unit Price"],
        ["Lamp", "Acme", "A", 10], ["Lamp", "Acme", "B", 20],
    ]), "p.xlsx")
    assert {(ref["field"], ref["cell"]) for ref in products[0]["source_refs"]} == {
        ("title", "A3"), ("vendor", "B3"), ("sku", "C3"), ("price", "D3")
    }


def test_variants_range_anchors_first_row_and_sheet_for_price_and_sku():
    products = [{
        "title": "Lamp", "variants": [{"sku": "L-1", "price": 12}, {"sku": "L-2", "price": 13}],
        "source_refs": [{"field": "variants", "sheet": "Spring Catalog", "cell_range": "C2:D3"}],
    }]
    data = multisheet_workbook_bytes([
        ("Archive", [["Title", "Price", "SKU"], ["Lamp", 99, "L-1"]]),
        ("Spring Catalog", [["Title", "Price", "SKU"], ["Lamp", 12, "L-1"], ["Lamp", 13, "L-2"]]),
    ])
    enrich_spreadsheet_source_refs(products, data, "p.xlsx")
    refs = {ref["field"]: ref for ref in products[0]["source_refs"] if ref.get("field") in {"sku", "price"}}
    assert refs["price"] | {"value": "12"} == refs["price"]
    assert (refs["price"]["sheet"], refs["price"]["cell"]) == ("Spring Catalog", "B2")
    assert (refs["sku"]["sheet"], refs["sku"]["cell"]) == ("Spring Catalog", "C2")


def test_identical_products_are_resolved_on_their_own_anchored_sheets():
    data = multisheet_workbook_bytes([
        ("Retail", [["Title", "Price", "SKU"], ["Lamp", 12, "L-1"]]),
        ("Wholesale", [["Title", "Price", "SKU"], ["Lamp", 12, "L-1"]]),
    ])
    products = [
        {"title": "Lamp", "variants": [{"sku": "L-1", "price": 12}], "source_refs": [{"field": "title", "sheet": name, "cell": "A2"}]}
        for name in ("Retail", "Wholesale")
    ]
    enrich_spreadsheet_source_refs(products, data, "p.xlsx")
    assert [next(ref for ref in product["source_refs"] if ref.get("field") == "sku")["sheet"] for product in products] == ["Retail", "Wholesale"]


def test_apostrophe_sheet_qualified_anchor_is_preserved():
    data = multisheet_workbook_bytes([("Maker's Picks", [["Title", "SKU"], ["Lamp", "L-1"]])])
    products = [{
        "title": "Lamp", "variants": [{"sku": "L-1"}],
        "source_refs": [{"field": "variants", "cell_range": "'Maker''s Picks'!A2:B2"}],
    }]
    enrich_spreadsheet_source_refs(products, data, "p.xlsx")
    sku = next(ref for ref in products[0]["source_refs"] if ref.get("field") == "sku")
    assert (sku["sheet"], sku["cell"]) == ("Maker's Picks", "B2")


def test_conflicting_sheet_anchors_remove_invalid_sku_and_do_not_guess():
    data = multisheet_workbook_bytes([
        ("One", [["Title", "SKU"], ["Lamp", "L-1"]]),
        ("Two", [["Title", "SKU"], ["Lamp", "L-1"]]),
    ])
    products = [{
        "title": "Lamp", "variants": [{"sku": "L-1"}],
        "source_refs": [
            {"field": "title", "sheet": "One", "cell": "A2"},
            {"field": "variants", "sheet": "Two", "cell_range": "A2:B2"},
            {"field": "sku", "sheet": "One", "cell": "A2"},
        ],
    }]
    enrich_spreadsheet_source_refs(products, data, "p.xlsx")
    assert not any(ref.get("field") == "sku" for ref in products[0]["source_refs"])


def test_unanchored_duplicate_across_sheets_is_unresolved():
    data = multisheet_workbook_bytes([
        ("One", [["Title", "SKU"], ["Lamp", "L-1"]]),
        ("Two", [["Title", "SKU"], ["Lamp", "L-1"]]),
    ])
    products = [{"title": "Lamp", "variants": [{"sku": "L-1"}]}]
    enrich_spreadsheet_source_refs(products, data, "p.xlsx")
    assert products[0]["source_refs"] == []


def test_wrong_sku_ref_is_not_repaired_from_another_sheet():
    data = multisheet_workbook_bytes([
        ("One", [["Title", "SKU"], ["Lamp", "WRONG"]]),
        ("Two", [["Title", "SKU"], ["Lamp", "L-1"]]),
    ])
    products = [{
        "title": "Lamp", "variants": [{"sku": "L-1"}],
        "source_refs": [
            {"field": "title", "sheet": "One", "cell": "A2"},
            {"field": "sku", "sheet": "One", "cell": "B2"},
        ],
    }]
    enrich_spreadsheet_source_refs(products, data, "p.xlsx")
    assert not any(ref.get("field") == "sku" for ref in products[0]["source_refs"])
